import pandas as pd
from openpyxl import load_workbook
import os 
import sys
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import re
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import copy
from io import BytesIO
import dropbox

class NDASubjects():
    """Class to collect status for each 
    form to allow subjects to be filtered
    out from NDA upload."""

    def __init__(self,network):
        self.network = network
        self.list_of_dict = []
        self.combined_df_folder ='/data/predict1/data_from_nda/formqc/'
        self.completion_status_path = '/PHShome/ob001/anaconda3/new_forms_qc/QC/'
        self.absolute_path = '/PHShome/ob001/anaconda3/nda_filter/'
        self.predict_path = '/data/predict1/data_from_nda/form_status/'

        with open(f'{self.absolute_path}forms_per_timepoint.json', 'r') as file:
            self.forms_per_tp = json.load(file)

        with open(f'{self.absolute_path}unique_form_vars.json', 'r') as file:
            self.unique_form_variables = json.load(file)

        self.all_forms = list(self.unique_form_variables.keys()) 
        
        self.subject_info = {}

        self.screening_df =  pd.read_csv(\
        f'{self.combined_df_folder}combined-{network}-screening-day1to1.csv',\
        keep_default_na=False)

        self.prescient_entry_statuses = pd.read_csv(\
        f'{self.completion_status_path}combined_prescient_completion_status.csv',\
        keep_default_na=False)

        self.form_status_dictionary = {}
        self.form_status_list = []
        self.error_df = pd.read_excel(\
        (f'/PHShome/ob001/anaconda3/new_forms_qc/QC'
        f'/site_outputs/{self.network}/combined/{self.network}_Output.xlsx'),\
            keep_default_na=False)
        
        self.excluded_forms = []
        if self.network =='PRESCIENT':
            self.excluded_forms = \
            ['informed_consent_run_sheet',\
            'family_interview_for_genetic_studies_figs']


    def run_script(self):
        self.all_timepoints = self.initalize_timepoint_list()
        self.collect_subject_info()
        self.call_form_status_functions()
        self.format_output()
    
    def call_form_status_functions(self):
        """Loops through each timepoint
        and calls functions to collect
        their form statuses and save 
        them into an excel spreadsheet"""

        for timepoint in self.all_timepoints:
            if self.network == 'PRESCIENT':
                self.current_tp_entry_statuses = \
                copy.deepcopy(self.prescient_entry_statuses)
                self.current_tp_entry_statuses = \
                self.current_tp_entry_statuses[\
                (self.current_tp_entry_statuses[\
                'Form_Timepoint'] == timepoint)]
            timepoint_str =\
            timepoint.replace('screen',\
            'screening').replace('baseln','baseline')
            self.organize_form_output(timepoint_str,timepoint)

    def format_output(self):
        """
        Formats output into pandas 
        dataframe and saved it as an
        excel file.
        """

        for subject,form_info in\
        self.form_status_dictionary.items():
            final_output_row = {}
            final_output_row['subject'] = subject
            final_output_row['HC or CHR'] =\
            self.subject_info[subject]['cohort']
            final_output_row['current timepoint'] =\
            self.subject_info[subject]['current_timepoint']
            for form,status in form_info.items():
                final_output_row[form] = status
            self.form_status_list.append(final_output_row)
        self.form_status_df = pd.DataFrame(self.form_status_list)
        self.form_status_df.to_excel(\
        f'{self.absolute_path}form_status_tracker_{self.network}.xlsx',\
        index = False)
        self.reformat_excel(\
        f'{self.absolute_path}form_status_tracker_{self.network}.xlsx')
        self.save_to_dropbox()

    
    def initalize_timepoint_list(self):
        """Organizes every timepoint as a key
        in a dictionary

        Returns
        ------------
        output_per_timepoint: dictionary
        with every timepoint as a key
        """

        timepoint_list = ['screen','baseln']
        for x in range(1,13):
            timepoint_list.append('month'+f'{x}')
        timepoint_list.extend(['month18','month24'])

        return timepoint_list
    
    def collect_cohort_str(self,cohort_num):
        """Translates the value of the
        chrcrit_part variable into the 
        appropriate cohort

        Parameters 
        ------------
        cohort_num: value of chrcrit_part

        Returns
        -----------
        cohort: corresponding cohort
        """

        if cohort_num in [1,'1',1.0,'1.0']:
            cohort = 'CHR'
        elif cohort_num in [2,'2',2.0,'2.0']:
            cohort = 'HC'
        else:
            cohort = 'Unknown'
        return cohort
    
    def collect_subject_info(self):
        """Collects cohort and current
        timepoint for each subject"""

        for row in self.screening_df.itertuples():
            self.subject_info[row.subjectid] =\
            {'cohort':self.collect_cohort_str(row.chrcrit_part),\
             'current_timepoint': row.visit_status_string}

    def organize_form_output(self,timepoint_str,timepoint):
        """Calls functions to check for the status of each
        form and organizes it into a dictionary
        
        Parameters
        ------------
        timepoint_str: string used to label timepoint
        timepoint: current timepoint
        """

        current_tp_df = pd.read_csv(\
        (f'{self.combined_df_folder}combined'
        f'-{self.network}-{timepoint_str}-day1to1.csv'),\
        keep_default_na=False)
        for row in current_tp_df.itertuples():
            all_tp_forms = self.collect_all_tp_forms(timepoint)
            for form in all_tp_forms:
                self.form_status_dictionary.setdefault(row.subjectid,{})
                self.form_status_dictionary[row.subjectid].setdefault(\
                form+'_' +timepoint_str,'')
                if form in self.excluded_forms:
                    self.form_status_dictionary[\
                    row.subjectid][form+'_' +timepoint_str] = \
                    'Completion Status Cannot Be Determined'
                    self.check_for_errors(row.subjectid,\
                    timepoint,form,timepoint_str)
                    continue
                self.collect_form_status(form,row,timepoint_str,timepoint)
                if self.form_status_dictionary[row.subjectid][\
                form+'_' +timepoint_str] != 'Not Marked Complete':
                    self.check_for_errors(row.subjectid,\
                    timepoint,form,timepoint_str)

    def collect_all_tp_forms(self,timepoint):
        """Collects every form at
        the current timepoint collected
        for either CHR or HC subjects
        
        Parameters
        ------------
        timepoint: current timepoint

        Returns
        -----------
        all_tp_forms: list of
        every form at timepoint
        """

        all_tp_forms = []
        for cohort in ['hc','chr']:
            for form in self.forms_per_tp[cohort][timepoint]:
                if form not in all_tp_forms:
                    all_tp_forms.append(form)
                    
        return all_tp_forms
            

    def collect_form_status(self,form,row,timepoint_str,timepoint):
        """Checks if current form is complete, marked as missing, 
        or if the completion status cannot be determined
        
        Parameters 
        --------------
        form: current form
        row: current row of combined dataframe
        timepoint_str: str used to label timepoint
        timepoint: current_timepoint
        """

        if 'complete_variable' in self.unique_form_variables[form].keys():
            current_completion_status =\
            self.check_prescient_na_values(form,row,timepoint)
            if current_completion_status not in \
            [2,2.0,'2','2.0',3,3.0,'3','3.0',4,4.0,'4','4.0']:
                self.form_status_dictionary[\
                row.subjectid][form+'_' +timepoint_str] = 'Not Marked Complete'
            elif current_completion_status in [4,4.0,'4','4.0'] or\
            ('missing_variable' in self.unique_form_variables[form].keys()\
            and hasattr(row,self.unique_form_variables[form]['missing_variable'])\
            and getattr(row,self.unique_form_variables[\
            form]['missing_variable']) in [1,1.0,'1','1.0']):
                self.form_status_dictionary[row.subjectid][form+'_' +timepoint_str]\
                = 'Marked as Missing'
            elif current_completion_status in [3,3.0,'3','3.0']:
                self.form_status_dictionary[row.subjectid][form +'_' +timepoint_str]\
                = 'Marked as Missing'
        elif 'complete_variable' not in self.unique_form_variables[form].keys():
            self.form_status_dictionary[row.subjectid][form+'_' +timepoint_str]\
            = 'Completion Status Cannot Be Determined'

    def check_prescient_na_values(self,form,row,timepoint):
        """Pulls completion status from raw entry 
        statuses that Prescient sends. If
        Prescient is not currently being checked,
        then will just return the completion
        value from the combined df.
        
        Parameters 
        ----------
        form: current form
        row: current row of combined df
        timepoint: current timepoint

        Returns 
        -----------
        comp_status: entry status 
        """
        comp_status = 0
        if self.network == 'PRESCIENT':
            df = self.current_tp_entry_statuses
            form_name = form.replace('_fu_hc','_fu') 
            current_form = df[(df['Subject'] == row.subjectid)\
            & (df['Form_Timepoint'] == \
            timepoint) & (df['Form_Translation'] == form_name)]
            for compl_row in current_form.itertuples():
                comp_status = compl_row.Completion_Status
                
        else:
            comp_status = getattr(\
            row,self.unique_form_variables[form]['complete_variable'])

        return comp_status
        

    def collect_psychs_err_forms(self,subject):
        """Collects the psychs forms that correspond
        to the current subject's cohort
        
        Parameters 
        --------------
        subject: current subject being checked

        Returns
        ----------------
        err_form_list: list of psychs forms
        """

        if self.subject_info[subject]['cohort'] == 'HC':
            err_form_list = ['psychs_p1p8_fu_hc','psychs_p9ac32_fu_hc']
        elif self.subject_info[subject]['cohort'] == 'CHR':
            err_form_list = ['psychs_p1p8_fu','psychs_p9ac32_fu']
        else:
            # if cohort unknown, assigns all psychs forms
            err_form_list = ['psychs_p1p8_fu','psychs_p9ac32_fu',\
                             'psychs_p1p8_fu_hc','psychs_p9ac32_fu_hc'] 
        return err_form_list

        
    def check_for_errors(self,subject,timepoint,form,timepoint_str):
        """Checks if there are any errors in the Main Report
        of the error tracker for the current form.
        
        Parameters 
        ------------
        subject: current subject
        timepoint: current timepoint
        form: current form
        timepoint_str: str used for timepoint label
        """

        error_df = self.error_df[self.error_df['Subject'] == subject]
        error_df = self.error_df[self.error_df['Subject'] == subject]
        error_df.columns = error_df.columns.str.replace(' ','_')
        
        for row in error_df.itertuples():
            err_form = getattr(\
                row,'General_Flag').split(' :')[0].strip().replace(' ','_').lower()
            err_tp = getattr(row,'Timepoint')
            error_sub = getattr(row,'Subject')
            resolved = getattr(row,'Date_Resolved')
            err_form_list = [err_form]
            if err_form == 'psychs_p1p8_fu/psychs_p9ac32_fu':
                err_form_list = self.collect_psychs_err_forms(subject)
            for curr_err_form in err_form_list:
                if curr_err_form == form and \
                timepoint == err_tp and error_sub == subject\
                and getattr(row,'Date_Resolved') == ''\
                and getattr(row,'Manually_Marked_as_Resolved') == '':
                    self.form_status_dictionary[\
                        subject][form + '_' + timepoint_str] = \
                        'Form Contains Unresolved Errors'


    def reformat_excel(self, file_path):
        """Adds colors to excel output depending 
        on the value of each cell
        
        Parameters
        -------------
        file_path: path to excel file being reformatted
        """

        wb = openpyxl.load_workbook(file_path)
        color_codes = {'red':'c47f7a','green':'b9d9b4',\
                    'yellow':'e6ebbc','grey':'B1B1B1'}
        fill_colors = {}
        for color,code in color_codes.items():
            fill_colors[color] = PatternFill(start_color=code,\
            end_color=code, fill_type="solid")
        for sheet in wb:
            for column in sheet.iter_cols(min_row=2, min_col=4):
                for cell in column:
                    if cell.value in ['Not Marked Complete',\
                    'Form Contains Unresolved Errors']:
                        cell.fill = fill_colors['red']
                    elif cell.value in ['Marked as Missing']:
                        cell.fill = fill_colors['yellow']
                    elif cell.value in ['Completion Status Cannot Be Determined']:
                        cell.fill = fill_colors['grey']
                    else:
                        cell.fill = fill_colors['green']
                    cell.border = Border(top=Side(style='thin'),\
                    right=Side(style='thin'), bottom=Side(style='thin'),\
                    left=Side(style='thin'))
            sheet = self.adjust_width(sheet)
            sheet.sheet_view.showGridLines = True 
            sheet.freeze_panes = "D2" 
        wb.save(file_path)
        for path in [self.absolute_path,self.predict_path]:
            wb.save(f"{path}form_status_tracker_{self.network}.xlsx")
    
    def adjust_width(self,sheet):
        """
        Adjusts width of current
        excel sheet.

        Parameters
        --------------
        sheet: input sheet

        Returns
        --------------------
        sheet: sheet after width is adjusted
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.coordinate in sheet.merged_cells:
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
        return sheet


    def collect_dropbox_credentials(self):
        """reads dropbox credentials from
        JSON file"""

        with open(\
        f'{self.completion_status_path}dropbox_credentials.json', 'r') as file:
            json_data = json.load(file)
        APP_KEY = json_data['app_key']
        APP_SECRET = json_data['app_secret']
        REFRESH_TOKEN = json_data['refresh_token']

        dbx = dropbox.Dropbox(
            app_key = APP_KEY,
            app_secret = APP_SECRET,
            oauth2_refresh_token = REFRESH_TOKEN
        )

        return dbx 
    
    def save_to_dropbox(self):
        """Saves form status output
        to dropbox folder"""

        dbx = self.collect_dropbox_credentials()
        dropbox_path = f'/Apps/form_status_trackers/'
        with open( \
        f"{self.absolute_path}form_status_tracker_{self.network}.xlsx", 'rb') as f:
                dbx.files_upload(f.read(),\
                f'{dropbox_path}form_status_tracker_{self.network}.xlsx',\
                mode=dropbox.files.WriteMode.overwrite)


if __name__ == '__main__':
    NDASubjects('PRONET').run_script()
    NDASubjects('PRESCIENT').run_script()
