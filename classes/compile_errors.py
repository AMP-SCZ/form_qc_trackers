import pandas as pd
import datetime
import os
import openpyxl
from openpyxl.utils import range_boundaries,get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl import load_workbook

class CompileErrors():
    """Class to format errors
    into final output as they
    are detected"""

    def __init__(self,timepoint,variable_translation_dict):
        self.variable_translation_dict = variable_translation_dict
        self.error_dictionary = {}
        self.timepoint = timepoint
        self.twentyone_day_tracker = []

    def append_error(self,row, message,  variable, form,sheet_list = ['Main Report']):
        """Function to append errors to a dictionary and collect
         comments from the form that those errors appeared in.

        Parameters
        ------------------
        message: message that will appear on the output 
        variable: variable involved in the error 
        form: form currently being flagged
        sheet_list: Sheets on the final tracker that this error
        will appear on
        """

        for sheet in sheet_list:
            self.error_dictionary.setdefault(sheet,{})
            self.error_dictionary[sheet].setdefault(row.subjectid,{})
            self.error_dictionary[sheet][row.subjectid].setdefault(form,{})
            column_headers = {"Subject":row.subjectid,"Timepoint":self.timepoint,\
            "Subject's Current Timepoint":row.visit_status_string,\
            "General Flag":f"{form.replace('_',' ').title()} : 0 flags detected.","Specific Flags":[],\
            "Variable Translations":'',"Date Flag Detected":f"{datetime.date.today()}",\
            "Time since Flag Detected":'',"Date Resolved":"",\
            "Manually Marked as Resolved":"","Sent to Site":"","Additional Comments":""}

            for column_header,default_value in column_headers.items():
                self.error_dictionary[sheet][\
                row.subjectid][form].setdefault(column_header,default_value)

            if f"{variable} : {message}" not in self.error_dictionary[sheet][row.subjectid][\
            form]["Specific Flags"]:
                self.error_dictionary[sheet][row.subjectid][\
                form]["Specific Flags"].append(f"{variable} : {message}")
                spec_flag_list = self.error_dictionary[sheet][row.subjectid][form]["Specific Flags"]
                flag_str = "flags"
                if len(spec_flag_list) == 1:
                    flag_str = "flag"
                self.error_dictionary[sheet][row.subjectid][form]["General Flag"] =\
                f"{form.replace('_',' ').title()} : {len(spec_flag_list)} {flag_str} detected."

                self.error_dictionary[sheet][row.subjectid][form]["Variable Translations"]\
                =self.add_variable_translations(spec_flag_list)

    def add_variable_translations(self, flag_list):
        """Adds variable translations for any variable that is present
        in the error list.

        Parameters
        -----------------------------
        flag_list: list of errors for the current row
        """
        translation_list = []
        for flag in flag_list:
            for word in flag.split():
                if word in self.variable_translation_dict.keys():
                    translation_list.append(\
                    self.variable_translation_dict[word])

        return translation_list

    def reformat_dataframe(self,output):
        """Changes output to proper
        format to be easily converted into 
        dataframe
        
        Parameters
        --------------
        Output: dictionary of output
        """
        formatted_output = {}
        for sheet_name, tracker in output.items():
            formatted_output.setdefault(sheet_name,[])
            for subject,subject_data in tracker.items():
                for form, form_data in subject_data.items():
                    formatted_output[sheet_name].append(form_data)

        return formatted_output

    def create_twenty_one_day_tracker(self,absolute_path):
        """Creates tracker to warn sites
        if they are approaching or exceeding 21
        days since screening psychs"""

        self.twentyone_day_tracker = sorted(self.twentyone_day_tracker,\
        key=lambda x: int(str(x['time_since_screening_psychs']).split(' ')[0]))
        filename = f'{absolute_path}site_outputs/PRONET/combined/PRONET_Output.xlsx'
        xls = pd.ExcelFile(filename)
        sheetname = 'Twenty One Day Tracker'
        twentyone_day_tracker_df = pd.DataFrame(self.twentyone_day_tracker)
        if os.path.exists(filename) and sheetname\
        in pd.ExcelFile(filename).sheet_names:
            with pd.ExcelWriter(filename, mode='a', engine='openpyxl',\
            if_sheet_exists = 'replace') as writer:
                old_df = pd.read_excel(filename,sheet_name=sheetname)
                for old_row in old_df.itertuples():
                    for new_row in twentyone_day_tracker_df.itertuples():
                        if old_row.subject == new_row.subject and\
                        old_row.recent_baseline_assessment\
                        == new_row.recent_baseline_assessment:
                            if hasattr(old_row,'sent_to_site')\
                             and hasattr(old_row,'manually_resolved'):
                                twentyone_day_tracker_df.at[new_row.Index,\
                                'sent_to_site'] = old_row.sent_to_site
                                twentyone_day_tracker_df.at[new_row.Index,\
                                'manually_resolved'] = old_row.manually_resolved
                twentyone_day_tracker_df.to_excel(writer, sheet_name=sheetname, index=False)
            workbook = load_workbook(filename = filename)
            worksheet = workbook[sheetname]
            for column in worksheet.columns:
                column_letter = get_column_letter(column[0].column)
                worksheet.column_dimensions[column_letter].width= 30
            workbook.save(filename)
        elif os.path.exists(filename):
            with pd.ExcelWriter(filename, mode='a', engine='openpyxl') as writer:
                twentyone_day_tracker_df.to_excel(writer, sheet_name=sheetname, index=False)
