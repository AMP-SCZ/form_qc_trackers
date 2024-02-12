import sys
import os
sys.path.append(os.path.abspath(os.path.dirname(__file__)))
from classes.data_checks import DataChecks
from classes.cross_form_checks import CrossForms
from classes.compile_errors import CompileErrors
import pandas as pd
import re
import datetime
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, colors, Protection,Color
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import range_boundaries,get_column_letter
from openpyxl import load_workbook
import numpy as np
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import openpyxl

class FormChecks():
    """Additional checks added to variables
    that passed through the FormIteration filters"""

    def __init__(self,dataframe,timepoint, sheet_title,process_variables,compile_errors):
        self.process_variables = process_variables
        self.variable_info_dictionary = self.process_variables.variable_info_dictionary
        self.all_barcode_variables = self.process_variables.all_barcode_variables
        self.all_blood_id_variables = self.process_variables.all_blood_id_variables
        self.all_blood_position_variables = self.process_variables.all_blood_position_variables
        self.all_blood_volume_variables = self.process_variables.all_blood_volume_variables
        self.timepoint_variable_lists = self.process_variables.timepoint_variable_lists

        self.compile_errors = compile_errors
        self.screening_df = self.process_variables.screening_df
        self.missing_code_list = self.process_variables.missing_code_list
        self.ampscz_df = self.process_variables.ampscz_df

    def call_extra_checks(self,form,variable,prescient,current_report_list,timepoint_variable_lists):
        self.prescient = prescient
        self.current_report_list = current_report_list
        self.timepoint_variable_lists = timepoint_variable_lists
        """Function to call all of
        the additional form checks specified below"""
        self.form = form
        self.variable = variable

        self.oasis_additional_checks()
        self.cssr_additional_checks()
        #self.penn_data_check(self.row.subjectid)
        if not self.prescient:
            self.blood_form_check()
        self.guid_format_check()
        self.date_format_check(form)
        self.global_function_check()
        self.barcode_format_check()
        self.age_iq_check()
        #self.iq_conversion_check()
        if 'blood' in self.form\
        and self.prescient == False:
            self.check_blood_duplicates()
            self.blood_vol_check()
            self.check_blood_freeze()
            self.check_blood_dates()
            self.cbc_unit_checks()
        self.cbc_differential_check()
        self.inclusion_psychs_check()
        #self.mri_scanner_check()
        if not self.prescient:
            self.tbi_check() 
            if self.checkbox_check() == False:
                self.compile_errors.append_error(\
                self.row,'value is empty',\
                self.variable,form,self.current_report_list)

    def age_iq_check(self):
        """Makes sure age and
        IQ are in proper ranges"""

        for age_var in ['chrdemo_age_mos_chr',\
        'chrdemo_age_mos_hc','chrdemo_age_mos2']:
            if self.variable == age_var:
                if getattr(self.row,age_var) != '' and\
                getattr(self.row,age_var) not in self.missing_code_list:
                    age = float(getattr(self.row,age_var))/12 
                    if age <12 or age > 31:
                        self.compile_errors.append_error(self.row,f'Improper age ({age}).',self.variable,\
                            self.form,['Main Report'])
        for iq_var in ['chriq_fsiq','chrpreiq_standard_score']:
            if self.variable == iq_var:
                if getattr(self.row,iq_var) != '' and\
                getattr(self.row,iq_var) not in self.missing_code_list:
                    iq = float(getattr(self.row,iq_var))
                    if iq < 50:
                        self.compile_errors.append_error(self.row,f'IQ is less than 50 ({iq}).',\
                        self.variable,self.form,\
                        ['Main Report','Cognition Report'])


    def collect_twenty_one_day_rule_dates(self,row):
        """Collectes appropriate baseline
        dates to check the 21 day rule"""

        self.missing_baseline_dates = []
        self.baseline_date_list = []
        self.max_baseline_date_variable_list = []
        baseline_date_variables = []
        for form in self.timepoint_variable_lists['baseln']:
            if 'interview_date' in self.variable_info_dictionary[\
            'unique_form_variables'][form]\
            and self.variable_info_dictionary[\
            'unique_form_variables'][form]['interview_date']\
            not in self.process_variables.excluded_21day_dates:
                baseline_date_variables.append(\
                self.variable_info_dictionary[\
                'unique_form_variables'][form]['interview_date'])
        matching_rows = self.screening_df.loc[self.screening_df['subjectid'] ==\
        self.row.subjectid, 'chrpsychs_scr_interview_date']
        self.psychs_interview_date = matching_rows.iloc[0] if not matching_rows.empty else '' 
        if getattr(self.row,'subjectid') in self.variable_info_dictionary['included_subjects']\
        and self.psychs_interview_date not in (self.missing_code_list+['']):
            self.psychs_interview_date = self.psychs_interview_date.replace('/', '-')
            self.psychs_interview_date = self.convert_date_format(self.psychs_interview_date)
            self.psychs_interview_date = re.search('\d{4}(-|/)\d{1,2}(-|/)\d{1,2}',\
            self.psychs_interview_date.replace('/', '-'))
            self.psychs_interview_date = \
            datetime.datetime.strptime(self.psychs_interview_date.group(), '%Y-%m-%d')
            for x, variable in enumerate(baseline_date_variables):
                extracted_baseline_date = re.search('\d{4}(-|/)\d{1,2}(-|/)\d{1,2}',\
                getattr(self.row, variable).replace('/', '-'))
                if extracted_baseline_date and\
                getattr(self.row, variable) not in self.missing_code_list:
                    self.baseline_date_list.append(datetime.datetime.strptime(\
                    extracted_baseline_date.group(), '%Y-%m-%d'))
                    self.max_baseline_date_variable_list.append(variable)
                else:
                    self.missing_baseline_dates.append(variable)
                    continue
        else:
            return False

        return True

    def twenty_one_day_rule(self,row,timepoint_variable_lists):
        """Check for baseline psychs to see if
        they are properly following the 21 day rule."""
        self.row = row
        self.timepoint_variable_lists = timepoint_variable_lists

        if self.collect_twenty_one_day_rule_dates(row) == True:
            if not self.baseline_date_list:
                return ''  
            date_of_last_baseline_assess = max(self.baseline_date_list)
            time_between = date_of_last_baseline_assess - self.psychs_interview_date
            time_since = datetime.datetime.today() - self.psychs_interview_date
            self.psychs_interview_date =\
            datetime.datetime.strftime(self.psychs_interview_date, "%Y-%m-%d")
            max_baseline_date_index = self.baseline_date_list.index(max(self.baseline_date_list))
            max_baseline_date = max(self.baseline_date_list)
            max_baseline_date_variable = self.max_baseline_date_variable_list[max_baseline_date_index]
            filtered_missing_dates = []
            for form in self.timepoint_variable_lists['baseln']:
                if 'interview_date' in self.variable_info_dictionary['unique_form_variables'][form]:
                    if self.variable_info_dictionary['unique_form_variables'][form]['interview_date']\
                    == max_baseline_date_variable:
                        max_form = form
                    if 'missing_variable' in self.variable_info_dictionary['unique_form_variables'][form]:
                        missing_var = self.variable_info_dictionary[\
                            'unique_form_variables'][form]['missing_variable']
                        for var in self.missing_baseline_dates:
                            if var == self.variable_info_dictionary[\
                            'unique_form_variables'][form]['interview_date'] and\
                            hasattr(self.row,missing_var) and\
                            getattr(self.row,missing_var)\
                            not in [1,1.0,'1','1.0']:
                                filtered_missing_dates.append(var)
            final_baseline_date = (f"{max_form}, "
            f"{datetime.datetime.strftime(max_baseline_date, '%Y-%m-%d')}")
            self.missing_baseline_dates = filtered_missing_dates 
            self.append_twenty_one_day_tracker_row(row,\
            max_form,date_of_last_baseline_assess,time_since)
            if time_between.days > 21:
                return self.append_twenty_one_day_error(time_between,final_baseline_date)

        return ''

    def append_twenty_one_day_tracker_row(self,row,max_form,date_of_last_baseline_assess,time_since):
        """Defines columns of 21 day tracker and
        appends a row.

        Parameters 
        ------------
        max_form: latest baseline assessment 
        date_of_last_baseline_assess: date of latest baseline assess 
        time_since: time between screening psychs and last baseline assess
        """

        if self.row.visit_status_string in ['baseln'] and not any(x in ['2',2,2.0,'2.0'] for x in\
        [self.row.psychs_p1p8_fu_hc_complete,self.row.psychs_p1p8_fu_complete,\
        self.row.psychs_p9ac32_fu_hc_complete,self.row.psychs_p9ac32_fu_complete]) \
        and not any(x in [1,1.0,'1','1.0'] for x in [self.row.chrpsychs_fu_missing_fu,\
        self.row.hcpsychs_fu_missing_fu,self.row.chrpsychs_fu_missing_fu_2,\
        self.row.hcpsychs_fu_missing_fu_2]):
            self.compile_errors.twentyone_day_tracker.append({'subject':self.row.subjectid,\
                    'time_since_screening_psychs':str(time_since.days) +' days',\
                    'recent_baseline_assessment':max_form,\
                    'date_of_recent_baseline_assessment':date_of_last_baseline_assess,\
                    'screening_psychs_date':self.psychs_interview_date,\
                    'Current Timepoint':self.row.visit_status_string,'sent_to_site':'',\
                    'manually_resolved':''})

    def append_twenty_one_day_error(self,time_between,final_baseline_date):
        """If current subject did not follow the 21
        day rule, will append error to final output
        
        Parameters
        --------------
        time_between: time between screening psychs and last baseline assess
        final_baseline_date: most recent baseline assessment
        """

        if len(self.missing_baseline_dates) > 1:
            sing_or_plur_str = (f"(This calculation does not include {self.missing_baseline_dates},"\
            " as the forms are missing dates and not marked as missing).")
        elif len(self.missing_baseline_dates) == 1:
            sing_or_plur_str = (f"(This calculation does not include {self.missing_baseline_dates},"\
            " as the form is missing its date and not marked as missing).")
        else:
            sing_or_plur_str = ''   
        if hasattr(self.row,'chrpsychs_fu_missing_spec_fu') and\
        (getattr(self.row,'chrpsychs_fu_missing_spec_fu')\
        in ['M6'] or getattr(self.row,'hcpsychs_fu_missing_spec_fu')\
        in ['M6'] or getattr(self.row,'chrpsychs_fu_missing_spec_fu_2') in ['M6']\
        or getattr(self.row,'hcpsychs_fu_missing_spec_fu_2') in ['M6']):
            return (f"M6 clicked, but there are {time_between.days} (> 21) days between"\
            f" screening PSYCHS ({self.psychs_interview_date}) and most recently"\
            f" completed baseline visit component ({final_baseline_date}). ") + sing_or_plur_str  

    def convert_date_format(self,date_string):
        """Function to convert dates to a y-m-d format"""
        
        formats = ["%Y-%m-%d", "%Y-%d-%m", "%d-%m-%Y", "%m-%d-%Y"]
        for date_format in formats:
            try:
                date_string = date_string.replace('/','').split(' ')[0]
                date = datetime.datetime.strptime(date_string, date_format)
                return datetime.datetime.strftime(date, "%Y-%m-%d")
            except Exception as e:
                continue


    def convert_timepoint_str(self,timepoint):
        if timepoint == 'baseln':
            timepoint_str = 'baseline'
        elif timepoint == 'screen':
            timepoint_str = 'screening'
        else:
            timepoint_str = timepoint
        return timepoint_str

    def date_format_check(self,form):
        """Check if dates are in proper format

        Parameters 
        ---------------
        form: current form in iteration
        """

        pattern = r"(\d+)(-|/)(\d+)(-|/)(\d+)"
        if self.variable in self.variable_info_dictionary['all_date_variables']\
        and hasattr(self.row,self.variable):
            if getattr(self.row,self.variable) not in ['','-3','nan','-9',-3,-9,-3.0,-9.0,\
            '-3.0','-9.0','1909-09-09','1903-03-03','1901-01-01']:
                date = getattr(self.row,self.variable)
                matches = re.findall(pattern,date)
                if not matches:
                    self.compile_errors.append_error(self.row,\
                    f" Date is in improper format. It was written as {date}",\
                    self.variable,form,self.current_report_list)
                else:
                    try:
                        if datetime.datetime.strptime(\
                        getattr(self.row,self.variable).split(' ')[0].replace('/','-'),'%Y-%m-%d')\
                        < datetime.datetime.strptime('2022-01-01','%Y-%m-%d'):
                            self.compile_errors.append_error(self.row,\
                            f" Date ({date}) is before January 2022",\
                            self.variable,form,self.current_report_list)
                        elif datetime.datetime.strptime(getattr(\
                        self.row,self.variable).split(' ')[0].replace(\
                        '/','-'),'%Y-%m-%d') > datetime.datetime.today():
                            self.compile_errors.append_error(self.row,\
                            f" Date ({date}) is in the future",\
                            self.variable,form,self.current_report_list)
                    except Exception as e:
                        print(e)
            elif getattr(self.row,self.variable) in \
            ['-3','-9',-3,-9,-3.0,-9.0,'-3.0','-9.0',\
            '1909-09-09','1903-03-03','1901-01-01']\
            and ('interview_date' in self.variable or\
            'chrmri_entry_date' in self.variable):
                self.compile_errors.append_error(self.row,\
                (f"Date is marked with a missing code ({getattr(self.row,self.variable)}),"
                " but all interview dates are required"),\
                self.variable,form,self.current_report_list)

    def guid_format_check(self):
        """Checks if GUID is in proper format"""

        if self.variable == 'chrguid_guid':
            if getattr(self.row,self.variable) != ''\
            and not re.search(r"^NDA[A-Z0-9]+$",\
            getattr(self.row,self.variable)):
                self.compile_errors.append_error(self.row,\
                    f'GUID in incorrect format. GUID was reported to be {getattr(self.row,self.variable)}.',\
                    self.variable,self.form,['Main Report'])

    def checkbox_check(self):
        """checks if any of the possible 
        checkboxes for current question were checked"""

        for main_variable, options in\
        self.process_variables.checkbox_variable_dictionary.items():
            if main_variable == self.variable:
                for item in options:
                    if getattr(self.row,item)\
                    not in [0,0.0,'0.0','0']: 
                        return True
                return False
        return True


    def mri_scanner_check(self):
        if self.variable == 'chrmri_scanner':
            self.mri_scanner_dict.setdefault(self.row.subjectid[:2],[])
            self.mri_scanner_dict[self.row.subjectid[:2]].append(self.row.chrmri_scanner)


    def blood_form_check(self):
        """Checks if time slept is
        not between 0 and 20 hours"""

        if self.variable == 'chrchs_timeslept':
            try:
                if not (0 <= float(getattr(self.row,'chrchs_timeslept')) < 20):
                    self.compile_errors.append_error(\
                    self.row,(f"Error in recorded time slept from the Current Status Form"\
                    f" (Recorded as {getattr(self.row,'chrchs_timeslept')})."),\
                    self.variable,self.form,['Main Report'])
            except Exception as e:
                print(e)
            try:  
                fasting_time = float(getattr(self.row,'time_fasting'))
                if not (0 <= fasting_time < 40): # change to 4 and 12 or 3 and 20
                    self.compile_errors.append_error(\
                    self.row,(f"Error in time fasting ({getattr(self.row,'time_fasting')})– either error"\
                    f" in blood draw date ({getattr(self.row,'chrblood_drawdate')})"\
                    f" in blood sample form or last time they ate ({getattr(self.row,'chrchs_ate')})"\
                    f" in current health status form."),'time_fasting',\
                    self.form,['Main Report','Blood Report'])
            except Exception as e:
                print(e)

    def blood_vol_check(self):
        if self.variable in self.all_blood_volume_variables:
            vol = getattr(self.row,self.variable)
            if vol not in (self.missing_code_list+['']) and\
            float(vol) <= 0:
                self.compile_errors.append_error(self.row,\
                f'Volume ({self.variable} = {vol}) is less then or equal to 0',\
                self.variable,self.form,['Main Report','Blood Report'])

    def check_blood_freeze(self):
        if self.variable == 'chrblood_wbfrztime':
            for proc_time_var in ['chrblood_wholeblood_freeze',\
            'chrblood_serum_freeze','chrblood_plasma_freeze']:
                proc_time =getattr(self.row, proc_time_var)
                if proc_time not in (self.missing_code_list+['']) and float(proc_time) > 480:
                    self.compile_errors.append_error(self.row,\
                    f'Processing time ({self.variable} = {proc_time}) is greater than 480',\
                    self.variable,self.form,['Main Report','Blood Report'])


    def check_blood_duplicates(self):
        """Checks for duplicate blood positions 
        and IDs"""

        filtered_columns = [col for col in self.ampscz_df.columns if 'blood' in col]
        filtered_columns.append('subjectid')
        self.filtered_blood_df = self.ampscz_df[filtered_columns]
        if self.variable =='chrblood_rack_barcode':
            blood_df = self.filtered_blood_df[self.filtered_blood_df[\
            'chrblood_rack_barcode']==getattr(self.row,'chrblood_rack_barcode')]
            for blood_pos_var in self.all_blood_position_variables:
                blood_df = blood_df[blood_df[blood_pos_var]==getattr(self.row,blood_pos_var)]
                for blood_row in blood_df.itertuples():
                    if blood_row.subjectid != self.row.subjectid:
                        if getattr(blood_row,blood_pos_var) == getattr(self.row,blood_pos_var)\
                        and getattr(blood_row,blood_pos_var) not in (self.missing_code_list+[''])\
                        and getattr(blood_row,'chrblood_rack_barcode') not in (self.missing_code_list+['']):
                            self.compile_errors.append_error(self.row,(\
                            f"Duplicate positions found in two different subjects ({self.row.subjectid} and {blood_row.subjectid} "\
                            f"both have {blood_pos_var} equal to {getattr(blood_row,blood_pos_var)})"\
                            f" and barcode equal to {self.row.chrblood_rack_barcode}."),\
                            blood_pos_var,self.form,['Blood Report'])
        if self.variable in self.all_blood_id_variables:
            blood_df = self.filtered_blood_df[\
            self.filtered_blood_df[self.variable]==getattr(self.row,self.variable)]
            for blood_row in blood_df.itertuples():
                if blood_row.subjectid != self.row.subjectid and\
                getattr(self.row,self.variable) not in (self.missing_code_list + ['']):
                    self.compile_errors.append_error(self.row,\
                    (f"Duplicate blood ID ({self.variable} = {getattr(self.row,self.variable)})",\
                    f" found between subjects {self.row.subjectid} and {blood_row.subjectid}."),\
                    self.variable,self.form,['Main Report','Blood Report'])
        elif self.variable in self.all_blood_volume_variables:
            if getattr(self.row,self.variable) not in (self.missing_code_list + [''])\
            and float(getattr(self.row,self.variable)) > 1.1:
                self.compile_errors.append_error(self.row,\
                f"Blood volume ({getattr(self.row,self.variable)}) is greater than 1.1.",\
                self.variable,self.form,['Main Report','Blood Report'])

    def check_blood_dates(self):
        """Function to check if the blood
        draw date is later than the date
        sent to lab."""

        if self.variable == 'chrblood_drawdate':
            if any(date  in (self.missing_code_list +['']) for\
            date in [self.row.chrblood_drawdate,self.row.chrblood_labdate]):
                return ''
            if datetime.datetime.strptime(\
            self.row.chrblood_drawdate,"%Y-%m-%d %H:%M")\
            > datetime.datetime.strptime(\
            self.row.chrblood_labdate,"%Y-%m-%d %H:%M"):
                self.compile_errors.append_error(self.row,\
                f"Blood draw date is later than date sent to lab.",\
                self.variable,self.form,['Main Report','Blood Report'])


    def barcode_format_check(self):
        """Makes sure blood barcodes are in
        proper format"""

        if self.variable in self.all_barcode_variables\
        and 'blood' in self.variable:
            if hasattr(self.row,self.variable):
                barcode = getattr(self.row,self.variable)
                if barcode not in (self.missing_code_list + [''])\
                and 'pronet' not in barcode.lower(): 
                    if len(barcode) < 10:
                        self.compile_errors.append_error(self.row,\
                        f"Barcode ({barcode}) length is less than 10 characters.",\
                        self.variable,self.form,['Blood Report'])
                    if any(not char.isdigit() for char in barcode):
                        self.compile_errors.append_error(self.row,\
                        f"Barcode ({barcode}) contains non-numeric characters.",\
                        self.variable,self.form,['Blood Report'])

    def call_scid_diagnosis_check(self,variable,row):
        self.variable = variable
        self.row = row
        for checked_variable,conditions in self.process_variables.scid_diagnosis_check_dictionary.items(): 
            if self.variable == checked_variable: 
                self.scid_diagnosis_check(\
                self.form,conditions['diagnosis_variables'],\
                conditions['disorder'],True,conditions['extra_conditionals'])
                self.scid_diagnosis_check(self.form,conditions[\
                'diagnosis_variables'],conditions['disorder'],\
                False,conditions['extra_conditionals'])
        self.scid_additional_checks(self.row,self.variable)


    def scid_additional_checks(self,row,variable):
        if variable == 'chrscid_a48_1':
            if row.chrscid_a27 in [3,3.0,'3','3.0'] and row.chrscid_a28 in [3,3.0,'3','3.0'] and\
               (row.chrscid_a48_1 not in [2,2.0,'2','2.0']): 
                self.compile_errors.append_error(self.row,f'Fulfills both main criteria but was counted incorrectly, a27, a28, a48_1',\
                self.variable,self.form,['Scid Report'])
            elif ((row.chrscid_a27 in [3,3.0,'3','3.0'] and (row.chrscid_a28 in\
            [2,2.0,'2','2.0'] or row.chrscid_a28 in [1,1.0,'1','1.0'])) or\
                 (row.chrscid_a28 in [3,3.0,'3','3.0'] and (row.chrscid_a27\
                in [2,2.0,'2','2.0'] or row.chrscid_a27 in [1,1.0,'1','1.0']))) and\
                 (row.chrscid_a48_1 not in [1,1.0,'1','1.0']):
                self.compile_errors.append_error(self.row,'Fulfills main criteria but further value was wrong, a27, a28, a48_1',\
                    self.variable,self.form,['Scid Report'])
        elif variable == 'chrscid_a51' and \
        row.chrscid_a26_53 not in (self.missing_code_list+['']):
            if float(row.chrscid_a26_53) <1 and (row.chrscid_a25\
            in [3,3.0,'3','3.0'] or row.chrscid_a51 in [3,3.0,'3','3.0']):
                self.compile_errors.append_error(self.row,('has no indication of total mde episodes'
                ' fulfilled in life even though fulfills current major depression. a26_53, a51'),\
                    self.variable,self.form,['Scid Report'])
            if float(row.chrscid_a26_53) > 0 and (row.chrscid_a25 not in [3,3.0,'3','3.0']\
            and row.chrscid_a51 not in [3,3.0,'3','3.0']):
                self.compile_errors.append_error(self.row,('fulfills more manic episodes than 0 but there'
                ' is no indication of past or current depressive episode. a26_53, a25'),\
                    self.variable,self.form,['Scid Report'])
        elif variable == 'chrscid_a1':
            if row.chrscid_a1 in [3,3.0,'3','3.0'] and\
            row.chrscid_a2 in [3,3.0,'3','3.0'] and\
            (row.chrscid_a22_1 not in [2,2.0,'2','2.0']):
                self.compile_errors.append_error(self.row,(f"Fulfills both main criteria"\
                " but was counted incorrectly, check a1, a2, a22_1"),\
                self.variable,self.form,['Scid Report'])
            if ((row.chrscid_a1 in [3,3.0,'3','3.0'] and (row.chrscid_a2 in [2,2.0,'2','2.0']\
            or row.chrscid_a2 in [1,1.0,'1','1.0'])) or\
                 (row.chrscid_a2 in  [3,3.0,'3','3.0'] and (row.chrscid_a1 in [2,2.0,'2','2.0']\
                 or row.chrscid_a1 in [1,1.0,'1','1.0']))) and\
                 (row.chrscid_a22_1 not in [1,1.0,'1','1.0']):
                self.compile_errors.append_error(self.row,'Fulfills main criteria but further value was wrong, a1, a2, a22_1',\
                self.variable,self.form,['Scid Report'])
        elif variable == 'chrscid_a25' and row.chrscid_a22 not\
        in (self.missing_code_list+['']) and row.chrscid_a22_1 not in (self.missing_code_list+['']):
            if float(row.chrscid_a22) > 4 and float(row.chrscid_a22_1) > 0  and (row.chrscid_a25\
            == '' or row.chrscid_a25 in self.missing_code_list):
                self.compile_errors.append_error(self.row,('A. MOOD EPISODES: subject fulfills more than 4'
                ' criteria of depression but further questions are not asked: start checking a22, a22_1, a25'),\
                    self.variable,self.form,['Scid Report'])


    def scid_diagnosis_check(self,form,conditional_variables,
                             disorder,fulfilled,extra_conditionals):
        try:
            row = self.row
            if fulfilled == True:
                for condition in conditional_variables:
                    if hasattr(self.row,condition) and \
                    getattr(self.row,condition) not in [3,3.0,'3','3.0']:
                        return ''
                if extra_conditionals != '':
                    for conditional in extra_conditionals:
                        if not eval(conditional):
                            return ''
                if getattr(self.row,self.variable) not in [3,3.0,'3','3.0']:
                    self.compile_errors.append_error(self.row,\
                    f'{disorder} criteria are fulfilled, but it is not indicated.',\
                    self.variable,form,['Scid Report'])
            else:
                for condition in conditional_variables:
                    if hasattr(self.row,condition) and \
                    getattr(self.row,condition) not in [3,3.0,'3','3.0']:
                        if hasattr(self.row,self.variable) and\
                        getattr(self.row,self.variable) in [3,3.0,'3','3.0']:
                            self.compile_errors.append_error(self.row,\
                            f'{disorder} criteria are NOT fulfilled, but it is indicated.',\
                            self.variable,form,['Scid Report'])
                if extra_conditionals != '':
                    for conditional in extra_conditionals:
                        if not eval(conditional):
                            if hasattr(self.row,self.variable) and\
                            getattr(self.row,self.variable) in [3,3.0,'3','3.0']:
                                self.compile_errors.append_error(self.row,\
                                f'{disorder} criteria are NOT fulfilled, but it is indicated.',\
                                self.variable,form,['Scid Report'])
        except Exception as e:
            print(e)

    def cssr_additional_checks(self):
        """Checks for contradictions in cssr form"""

        lifetime_pastyear_dict = {'chrcssrsb_cssrs_actual':'chrcssrsb_sb1l',\
        'chrcssrsb_cssrs_nssi':'chrcssrsb_sbnssibl','chrcssrsb_cssrs_yrs_ia':'chrcssrsb_sb3l',\
        'chrcssrsb_cssrs_yrs_aa':'chrcssrsb_sb4l',\
        'chrcssrsb_cssrs_yrs_pab':'chrcssrsb_sb5l','chrcssrsb_cssrs_yrs_sb':'chrcssrsb_sb6l'}
        lifetime_pastyear_greater_dict = {'chrcssrsb_idintsvl':'chrcssrsb_css_sipmms',\
        'chrcssrsb_snmacatl':'chrcssrsb_cssrs_num_attempt',\
        'chrcssrsb_nminatl':'chrcssrsb_cssrs_yrs_nia','chrcssrsb_nmabatl':'chrcssrsb_cssrs_yrs_naa'}
        for x in range(1,6):
            if self.variable == 'chrcssrsb_si{x}l':
                if getattr(self.row, f'chrcssrsb_si{x}l') in [2,2.0,'2','2.0']\
                and getattr(self.row, f'chrcssrsb_css_sim{x}') in [1,1.0,'1','1.0']:
                    self.compile_errors.append_error(self.row,f'Past month does not equal lifetime.',\
                    self.variable,self.form,['Main Report'])
        for key_past_months, value_lifetime in lifetime_pastyear_greater_dict.items():
            if self.variable == key_past_months:
                try:
                    if getattr(self.row,key_past_months) not in\
                    self.missing_code_list and \
                    getattr(self.row,value_lifetime) not in self.missing_code_list\
                    and float(getattr(self.row,key_past_months))\
                    < float(getattr(self.row,value_lifetime)):
                        self.compile_errors.append_error(self.row,(f'Lifetime ({getattr(self.row,key_past_months)}) cannot',\
                        f'be lower than past month ({getattr(self.row,value_lifetime)}).'),\
                        self.variable,self.form,['Main Report'])
                except Exception as e:
                    print(e)
        for key_past_months,value_lifetime in lifetime_pastyear_dict.items():
            if self.variable == key_past_months:
                if getattr(self.row,key_past_months) in [1,1.0,'1','1.0']\
                and getattr(self.row,value_lifetime) in [2,2.0,'2','2.0']:
                    self.compile_errors.append_error(self.row,f'Lifetime value cannot be different that past month value.',\
                    self.variable,self.form,['Main Report'])

    def inclusion_psychs_check(self):
        """Checks for contradictions between
        inclusion/exclusion form and psychs"""

        if self.variable == 'chrcrit_part':
            if (self.row.chrcrit_part in [1,1.0,'1','1.0']\
            and self.row.chrcrit_inc3 in [0,0.0,'0','0.0']):
                self.compile_errors.append_error(self.row,\
                f'CHR subject does not fulfill CHR-criteria on PSYCHS.',\
                self.variable,self.form,['Main Report'])
            elif (self.row.chrcrit_part in [2,2.0,'2','2.0']\
            and self.row.chrcrit_inc3 in [1,1.0,'1','1.0']):
                self.compile_errors.append_error(self.row,\
                f'HC subject fulfills CHR-criteria on PSYCHS.',\
                self.variable,self.form,['Main Report'])

    def oasis_additional_checks(self):
        """Checks for contradictions in OASIS form"""

        try:
            if self.variable == f'chroasis_oasis_1':
                for x in range(2,6):
                    if getattr(self.row,f'chroasis_oasis_{x}') not in self.missing_code_list\
                    and float(getattr(self.row,f'chroasis_oasis_{x}')) > 0 and\
                    self.row.chroasis_oasis_1 in [0,0.0,'0','0.0']:
                        self.compile_errors.append_error(self.row,(f'No anxiety at all (last week) cannot have anxiety '\
                        f'level or be influenced by anxiety (last week) (chroasis_oasis_{x}).'),\
                        self.variable,self.form,['Main Report'])
            if self.variable == 'chroasis_oasis_3':
                if self.row.chroasis_oasis_3 not in\
                self.missing_code_list and self.row.chroasis_oasis_4 not in self.missing_code_list\
                and self.row.chroasis_oasis_5 not in self.missing_code_list and \
                float(self.row.chroasis_oasis_3) <2 and \
                (float(self.row.chroasis_oasis_4 > 1) or float(self.row.chroasis_oasis_5)>0):
                    self.compile_errors.append_error(self.row,(f'If lifestyle is not affected (chroasis_oasis_3<2) no'\
                    f'lifestyle situation can be described to be affected (chroasis_oasis_4 and chroasis_oasis_5)'),\
                    self.variable,self.form,['Main Report'])
        except Exception as e:
            print(e)

    def find_iq_age(self):
        """Finds age at the time of the iq
        assessment"""

        try:
            demographics_date = datetime.datetime.strptime(\
            self.row.chrdemo_interview_date, "%Y-%m-%d") 
            iq_date = datetime.datetime.strptime(\
            self.row.chriq_interview_date, "%Y-%m-%d")  
            days_between = (iq_date - demographics_date).days
            months_between = int(days_between/30)
            return months_between
        except Exception as e:
            print(e)
            return 0

    def find_days_between(self,d1,d2):
        """finds the days between two dates
        
        Parameters
        -----------------
        d1: first date
        d2: second date
        """

        date_format = "%Y-%m-%d"
        date1 = datetime.datetime.strptime(\
        d1.split(' ')[0], date_format)
        date2 = datetime.datetime.strptime(\
        d2.split(' ')[0], date_format)
        date_difference = date2 - date1
        days_between = date_difference.days

        return abs(days_between)

    def check_if_complete_and_not_missing(self,form):
        """Checks if a form is marked as complete
        and not marked as missing

        Parameters
        ----------------
        form: forms being checked
        """

        complete_var = self.variable_info_dictionary[\
        'unique_form_variables'][form]['complete_variable']
        if getattr(self.row,complete_var) in [2,2.0,'2','2.0']:
            if 'missing_variable' in self.variable_info_dictionary[\
            'unique_form_variables'][form]:
                missing_var = self.variable_info_dictionary[\
                'unique_form_variables'][form]['missing_variable']
            else:
                return True
            if hasattr(self.row,\
            self.variable_info_dictionary[\
            'unique_form_variables'][form]['missing_variable'])\
            and str(getattr(self.row,\
            self.variable_info_dictionary[\
            'unique_form_variables'][form]['missing_variable']))\
            in ['','nan','False', "0", "0.0" ,"'0'","'0.0'"]:
                return True
            else:
                return False

    def collect_age(self):
        """Collects current age of 
        subject by checking each possible
        age variable"""

        age = ''
        for age_var in ['chrdemo_age_mos_chr',\
        'chrdemo_age_mos_hc','chrdemo_age_mos2']:
            if hasattr(self.row,age_var)\
            and getattr(self.row,age_var)\
            not in (self.missing_code_list+['']):
                age = int(getattr(self.row,age_var))
                age = age - self.find_iq_age()
                break
        return age

    def age_range_workaround(self,age):
        """Workaround for ages that fall on the 
        boundaries of a range and may fall 
        in one or another depending on how
        it is rounded

        Parameters
        --------------
        age: age being checked
        """

        current_age_ranges = []
        for age_range in self.all_iq_age_ranges:
            if (age == age_range[-1] or\
            age == age_range[-2]) and age < 360:
                possibly_next_range = True
                current_age_ranges.append(age_range)
                current_age_ranges.append(self.all_iq_age_ranges[\
                self.all_iq_age_ranges.index(age_range) + 1])
                flag_error = False
            elif (age == age_range[0]\
        or age == age_range[1]) and age > 195:
                possibly_previous_range = True
                current_age_ranges.append(age_range)
                current_age_ranges.append(self.all_iq_age_ranges[\
                self.all_iq_age_ranges.index(age_range) - 1])
                flag_error = False
            elif age in age_range:
                current_age_ranges.append(age_range)
                flag_error = True
        return current_age_ranges, flag_error

    def fsiq_check(self):
        """Checks if FSIQ conversion was done
        properly"""

        if self.variable == 'chriq_fsiq':
            for fsiq_row in self.fsiq_conversion_df.itertuples():
                if str(fsiq_row.t_score).replace(' ','')\
                == str(self.row.chriq_tscore_sum).replace(' ',''):
                    if str(fsiq_row.fsiq).replace(' ','')\
                    != str(self.row.chriq_fsiq).replace(' ',''):
                        self.compile_errors.append_error(self.row,(f'FSIQ-2 Conversion not done properly.'
                        f'Entered value was {self.row.chriq_fsiq}, but should be {fsiq_row.fsiq}'),\
                        self.variable,self.form,['Main Report','Cognition Report'])

    def loop_iq_table(self,age,conversion_col,t_score_col,iq_variable):
        """Loops through IQ table to make sure
        conversions were done properly

        Parameters
        -----------
        age: age of current subject
        conversion_col: current column being converted
        t_scor_col: column of the T-scores 
        iq_variable: variable associated with conversion
        """

        current_age_ranges,flag_error = self.age_range_workaround(age)
        for range_index in range(0,len(current_age_ranges)):
            any_match = False
            columns_to_keep = self.iq_conversion_df.columns[self.iq_conversion_df.iloc[0].apply(\
                                lambda x: current_age_ranges[range_index] == x)]
            iq_df = self.iq_conversion_df[columns_to_keep].copy()
            iq_df_filtered = iq_df.iloc[1:].copy().reset_index(drop=True)
            iq_df_filtered.columns = iq_df_filtered.iloc[0]
            for iq_row in iq_df_filtered.itertuples():
                if str(getattr(self.row,iq_variable)).replace(' ','') in\
                self.convert_range_to_list(str(getattr(iq_row,conversion_col)),True):
                    if str(iq_row.t_score).replace(' ','')\
                    != str(getattr(self.row,t_score_col)).replace(' ',''):
                        print(getattr(self.row,t_score_col))
                        if len(current_age_ranges) == 2 and\
                        range_index ==1 and not any_match:
                            flag_error = True
                        if current_age_ranges[range_index]\
                        == current_age_ranges[-1] and flag_error == True:
                            if len(current_age_ranges) == 1:
                                error_message = (f'T-Score Conversion not done properly.'
                                f'Entered value was {getattr(self.row,t_score_col)}, but should be {iq_row.t_score}')
                            else:
                                error_message = (f'T-Score Conversion not done properly.'
                                f'Entered value was {getattr(self.row,t_score_col)}' 
                                f'(cannot calculate proper value that it should be due to insufficient age rounding).')
                            self.compile_errors.append_error(\
                            self.row,error_message,self.variable,\
                            self.form,['Main Report','Cognition Report'])
                    else:
                        any_match = True

    def iq_conversion_check(self):
        """Loops through different IQ
        variables being converted and 
        calls functions to check them"""

        for iq_variable in ['chriq_vocab_raw','chriq_matrix_raw']:
            possibly_next_range = False
            possibly_previous_range = False
            if self.variable == iq_variable and \
            self.row.chriq_assessment in [1,1.0,'1','1.0']:
                if iq_variable == 'chriq_vocab_raw':
                    conversion_col = 'vc'
                    t_score_col = 'chriq_tscore_vocab'
                else:
                    conversion_col = 'mr'
                    t_score_col = 'chriq_tscore_matrix'
                age = self.collect_age()

                if age !='' and age>191 and self.row.chriq_tscore_vocab != ''\
                and self.row.chriq_vocab_raw not in self.missing_code_list:
                    self.loop_iq_table(age,conversion_col,t_score_col,iq_variable)
        self.fsiq_check()


    def convert_range_to_list(self,range_str,str_conv = False):
        """Converts string with number range to a list of 
        the numbers included in that range. Used for IQ age checks.

        Parameters
        -------------
        range_str: string of number range
        str_conv:
        """
        
        range_list = []
        if '-' not in range_str:
            if str_conv ==True:
                return [str(range_str).replace(' ','')]
            else:
                return range_str
        first_item = int(range_str.split('-')[0])
        last_item = int(range_str.split('-')[1])
        for x in range(first_item,last_item+1):
            if str_conv ==True:
                new_item = str(x).replace(' ','')
            else:
                new_item = x
            range_list.append(new_item)
        return range_list
         
    def global_function_check(self):
        """Checks for contradictions in the
        global functioning forma"""

        try:
            if self.variable == 'chrgfs_gf_social_low':
                if self.row.chrgfs_gf_social_low not in self.missing_code_list\
                and (((float(self.row.chrgfs_gf_social_low) \
                > float(self.row.chrgfs_gf_social_scale) and self.row.chrgfs_gf_social_scale\
                not in self.missing_code_list) or\
                ((float(self.row.chrgfs_gf_social_low) > float(self.row.chrgfs_gf_social_high)\
                and self.row.chrgfs_gf_social_high not in self.missing_code_list)))):
                    self.compile_errors.append_error\
                    (self.row,(f"Social Scale low score ({self.row.chrgfs_gf_social_low}) is not the lowest score"\
                     f"(current score = {self.row.chrgfs_gf_social_scale}, high score = {self.row.chrgfs_gf_social_high})."),\
                    self.variable,self.form,['Main Report'])
            elif self.variable == 'chrgfs_gf_social_scale':
                if self.row.chrgfs_gf_social_scale not in self.missing_code_list\
                and (float(self.row.chrgfs_gf_social_scale)\
                > float(self.row.chrgfs_gf_social_high) and\
                self.row.chrgfs_gf_social_high not in self.missing_code_list):
                    self.compile_errors.append_error\
                    (self.row,(f"Social Scale current score ({self.row.chrgfs_gf_social_scale}) is greater",\
                    f"than the high score ({self.row.chrgfs_gf_social_high})."),\
                    self.variable,self.form,['Main Report'])
            elif self.variable == 'chrgfs_gf_role_low':
                if self.row.chrgfs_gf_role_low not in self.missing_code_list\
                and ((float(self.row.chrgfs_gf_role_low)\
                > float(self.row.chrgfs_gf_role_scale) and\
                self.row.chrgfs_gf_role_scale not in self.missing_code_list)\
                or (float(self.row.chrgfs_gf_role_low) > float(self.row.chrgfs_gf_role_high)\
                and self.row.chrgfs_gf_role_high not in self.missing_code_list)):
                    self.compile_errors.append_error\
                    (self.row,(f"Role Scale low score ({self.row.chrgfs_gf_role_low}) is not the lowest score",\
                    f"(current score = {self.row.chrgfs_gf_role_scale}, high score = {self.row.chrgfs_gf_role_high})."),\
                    self.variable,self.form,['Main Report'])
            elif self.variable == 'chrgfs_gf_role_scale':
                if self.row.chrgfs_gf_role_scale not in self.missing_code_list\
                and float(self.row.chrgfs_gf_role_scale) > float(self.row.chrgfs_gf_role_high)\
                and self.row.chrgfs_gf_role_high not in self.missing_code_list:
                    self.compile_errors.append_error\
                    (self.row,(f"Role Scale current score ({self.row.chrgfs_gf_role_scale}) is greater",
                    f"than the high score ({self.row.chrgfs_gf_role_high})."),\
                    self.variable,self.form,['Main Report'])
        except Exception as e:
            print('----')
            print(e)

    def tbi_check(self):
        """Multiple specific checks for the TBI form"""

        if self.variable == 'chrtbi_severe_inj':
            injury_rating = self.row.chrtbi_severe_inj
            if isinstance(self.row.chrtbi_severe_inj, str):
                injury_rating = self.row.chrtbi_severe_inj.replace('.0','')
            if injury_rating!='' and int(injury_rating) > 6: # makes sure 
                self.compile_errors.append_error(self.row,'Most severe head injury > 6 but participant not excluded',\
                self.variable,self.form,['Main Report'])
        elif self.variable == 'chrtbi_sourceinfo':
            if self.row.chrtbi_sourceinfo in [3,3.0,'3','3.0']:
                if self.row.chrtbi_subject_head_injury != self.row.chrtbi_parent_headinjury:
                    self.compile_errors.append_error(self.row,\
                    ("Subject and parent answered differently to whether"
                    " or not the subject has ever had a head injury."),\
                    self.variable,self.form,['Main Report'])
            elif self.row.chrtbi_sourceinfo in [1,1.0,'1','1.0',2,2.0,'2','2.0']\
            and (self.row.chrtbi_subject_head_injury in\
            [1,1.0,'1','1.0',0,0.0,'0','0.0'] and\
            self.row.chrtbi_parent_headinjury in [1,1.0,'1','1.0',0,0.0,'0','0.0']):
                self.compile_errors.append_error(\
                self.row,("Subject and parent not both selected as source of information,"\
                " but answers appear to be provided by both the subject and parent."),\
                self.variable,self.form,['Secondary Report'])


    def cbc_differential_check(self):
        """Checks for additional errors in
        CBC with differntial form"""

        if self.row.chrcbc_missing in ['','nan','False', "0", "0.0" ,"'0'","'0.0'"]:
            if self.variable == 'chrcbc_wbcinrange':
                if not any(x in (self.missing_code_list+['',999,'999',999.0,'999.0'])\
                for x in [self.row.chrcbc_wbcsum,\
                self.row.chrcbc_wbc]):
                    wbc_sum = float(self.row.chrcbc_wbcsum)
                    wbc = float(self.row.chrcbc_wbc)
                    if (wbc_sum < (wbc-(0.1*wbc))) or wbc_sum > (wbc+(0.1*wbc)):
                        self.compile_errors.append_error(\
                            self.row,f'Sum of absolute neutrophils, lymphocytes,'
                            f' monocytes, basophils, and eosinophils'
                            f' ({self.row.chrcbc_wbcsum}) is not within 10% of '
                            f' WBC count ({self.row.chrcbc_wbc}).',self.variable,self.form,\
                            ['Blood Report'])
            elif self.variable == 'chrblood_cbc':
                if self.row.chrblood_cbc in [1,1.0,'1','1.0']\
                and getattr(self.row,'cbc_with_differential_complete') not in [2,2.0,'2.0','2'] and\
                self.row.chrblood_interview_date not in (self.missing_code_list+['']):
                    time_since_blood = self.find_days_between(\
                    str(self.row.chrblood_interview_date),str(datetime.datetime.today()))
                    if time_since_blood > 5:
                        self.compile_errors.append_error(\
                        self.row,('Blood form indicates EDTA tube was sent to lab for CBC'
                        f', but CBC form has not been completed.'),\
                        self.variable,self.form, ['Main Report','Blood Report'])
                elif self.row.chrblood_cbc not in [1,1.0,'1','1.0']\
                and getattr(self.row,'cbc_with_differential_complete') in [2,2.0,'2.0','2']\
                and self.row.chrcbc_missing not in ['1','1.0',1,1.0]:
                    self.compile_errors.append_error(\
                    self.row,('Blood form indicates EDTA tube was not sent to lab for CBC'
                    f', but CBC form has been completed and not marked as missing.'),\
                    self.variable,self.form, ['Main Report','Blood Report'])

    def cbc_unit_checks(self):
        gt_unit_ranges = {'chrcbc_wbc':30,'chrcbc_neut':20,'chrcbc_rbc':12,\
        'chrcbc_lymph':20,'chrcbc_eos':5,'chrcbc_monos':5,'chrcbc_baso':5}
        lt_unit_ranges = {'chrcbc_monos':0.1,'chrcbc_hct':1}
        for var,value_range in gt_unit_ranges.items():
            if self.variable == var and var not in (self.missing_code_list + ['']):
                if float(getattr(self.row,var)) > value_range:
                    self.compile_errors.append_error(self.row,\
                    f'Incorrect units used ({var} = {float(getattr(self.row,var))})',\
                    self.variable,self.form, ['Main Report','Blood Report'])
        for var,value_range in lt_unit_ranges.items():
            if self.variable == var and var not in (self.missing_code_list + ['']):
                if float(getattr(self.row,var)) < value_range:
                    self.compile_errors.append_error(self.row,\
                    f'Incorrect units used ({var} = {float(getattr(self.row,var))})',\
                    self.variable,self.form, ['Main Report','Blood Report'])

    def penn_data_check(self,subject):
        if 'penncnb' in self.timepoint_variable_lists[self.timepoint]\
        and self.variable == 'chrpenn_complete' and self.form =='penncnb':
            if not isinstance(self.penn_data_summary_df, pd.DataFrame):
                timepoint_str = self.convert_timepoint_str(self.timepoint)
                self.penn_data_summary_df = pd.read_csv(\
                (f'{self.penn_path}{self.combined_cognition_folder}combined-AMPSCZ'
                '-data_{timepoint_str.replace("month","month_")}-day1to1.csv'),\
                keep_default_na = False)
            if subject in self.variable_info_dictionary['included_subjects']:
                for penn_row in self.penn_data_summary_df.itertuples():
                    if penn_row.subject_id == subject:
                        if '-' in str(penn_row.cnb_data) and penn_row.cnb_data !=''\
                        and penn_row.cnb_protocol in [1,1.0,'1','1.0']\
                        and abs(int(penn_row.cnb_data)) > 5:
                            self.compile_errors.append_error(\
                            self.row,(f"Penn Data has been missing for {abs(int(penn_row.cnb_data))} days."\
                            "Please check to make sure subject ID is in the correct format."),\
                            'Penn Data','penncnb',['Main Report','Cognition Report'])


                                
