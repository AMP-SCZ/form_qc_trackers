import pandas as pd
import os
import sys
import json
import dropbox
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, colors, Protection,Color
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import range_boundaries,get_column_letter
from openpyxl import load_workbook,Workbook
import numpy as np
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import openpyxl
import dropbox
import webbrowser
import base64
import requests
import json
from io import BytesIO
import ast
from openpyxl.formatting.formatting import ConditionalFormattingList

parent_dir = "/".join(os.path.realpath(__file__).split("/")[0:-2])
sys.path.insert(1, parent_dir)

from utils.utils import Utils

class CreateTrackers():

    def __init__(self, formatted_col_names):
        self.utils = Utils()
        with open(f'{self.utils.absolute_path}/config.json','r') as file:
            self.config_info = json.load(file)
        self.output_path = self.config_info['paths']['output_path']
        if self.config_info["testing_enabled"] == "True":
            self.output_path += "testing/"
        
        self.all_reports = ['Main Report','Secondary Report']
        self.site_reports = ['Main Report']
        self.all_report_df = {}
        self.all_pronet_sites = self.utils.all_pronet_sites
        self.all_prescient_sites = self.utils.all_prescient_sites
        self.all_sites = self.utils.all_sites
        self.site_translations = self.utils.site_full_name_translations
        self.old_output_csv_path = f'{self.output_path}combined_outputs/old_output/combined_qc_flags.csv'
        self.new_output_csv_path = f'{self.output_path}combined_outputs/new_output/combined_qc_flags.csv'
        self.formatted_outputs_path = f'{self.output_path}formatted_outputs/'
        if not os.path.exists(self.formatted_outputs_path):
            os.makedirs(self.formatted_outputs_path)
        self.dropbox_output_path =  f'{self.output_path}formatted_outputs/dropbox_files/'
        self.colors = {"green" : PatternFill(start_color='b9d9b4', end_color='b9d9b4', fill_type='gray125'),
                       "yellow" : PatternFill(start_color='e8e7b0', end_color='e8e7b0', fill_type='gray125'),
                       "blue" : PatternFill(start_color='d8cffc', end_color='d8cffc', fill_type='gray125'),
                       "orange" : PatternFill(start_color='ebba83', end_color='ebba83', fill_type='gray125'),
                       "red" : PatternFill(start_color='de9590', end_color='de9590', fill_type='gray125'),
                       "grey" : PatternFill(start_color='ededed', end_color='ededed', fill_type='gray125')}
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),bottom=Side(style='thin'))
        self.formatted_column_names = formatted_col_names
        self.melbourne_ras = self.utils.load_dependency_json('melbourne_ra_subs.json')
        
    def run_script(self):
        self.combined_tracker = pd.read_csv(self.new_output_csv_path, keep_default_na= False)
        self.collect_new_reports()
        self.generate_reports()
        self.upload_trackers()

    def collect_new_reports(self):
        for row in self.combined_tracker.itertuples():
            if row.reports == '':
                continue
            reports = (row.reports).split(' | ')
            for report in reports:
                if report not in self.all_reports:
                    self.all_reports.append(report)

    def generate_reports(self):
        for network in ['PRONET','PRESCIENT']:
            for report in self.all_reports:
                network_df = self.combined_tracker[
                self.combined_tracker['network']==network]
                report_df = network_df[
                network_df['reports'].str.contains(report)]
                self.all_report_df[report] = report_df
                if report_df.empty:
                    continue
                report_df = self.convert_to_shared_format(report_df, network)
                if report in self.all_reports: 
                    print(network)
                    combined_path = f'{self.dropbox_output_path}{network}/combined/'
                    if not os.path.exists(combined_path):
                        os.makedirs(combined_path)
                    self.format_excl_sheet(report_df,report,
                    combined_path,
                    f'{network}_combined_Output.xlsx')
                    self.loop_sites(network, report, report_df)

    def loop_sites(self, network, report, report_df):
        for site_abr in self.all_sites[network]:
            if site_abr in self.utils.site_full_name_translations.keys():
                site = self.utils.site_full_name_translations[site_abr]
            else:
                site = site_abr
            print(site)
            if report != 'Main Report' and site_abr != 'ME':
                continue 
            if report != 'Non Team Forms' and site_abr == 'ME':
                continue 
            if site_abr == 'ME':
                self.loop_ras(network,site, report, report_df)
            site_path = f'{self.dropbox_output_path}{network}/{site}/'
            if not os.path.exists(site_path):
                os.makedirs(site_path)
            site_df = report_df[report_df['Participant'].str[:2] == site_abr]
            self.format_excl_sheet(site_df,
            report,site_path,
            f'{network}_{site_abr}_Output.xlsx')

    def loop_ras(self,network,site, report, report_df):
        for ra, subjects in self.melbourne_ras.items():
            print(ra)
            ra_path = f'{self.dropbox_output_path}{network}/{site}/{ra}/'
            ra_df = report_df[report_df['Participant'].isin(subjects)]
            self.format_excl_sheet(ra_df,
            report,ra_path,
            f'{network}_Melbourne_RA_Output.xlsx')

    def upload_trackers(self):
        fullpath = self.output_path + '/formatted_outputs/dropbox_files/'
        for root, dirs, files in os.walk(fullpath):
            for file in files:
                if file.endswith('Output.xlsx'):
                    #print(root + '/' + file)
                    full_path = root + '/' + file
                    local_path = root.replace(fullpath,'') + '/' + file
                    #print(root.replace(fullpath,'') + '/' + file)
                    self.save_to_dropbox(full_path,local_path)
                    print(local_path)

    def format_excl_sheet(self, df, report, folder, filename):
        full_path = folder + filename
        if not os.path.exists(folder):
            os.makedirs(folder)

        if not os.path.exists(folder + filename):

            df.to_excel(folder + filename, sheet_name = report, index = False)

        with pd.ExcelWriter(full_path, mode='a',\
        engine='openpyxl',if_sheet_exists = 'replace') as writer:                
            df.to_excel(writer, sheet_name=report, index=False)

        workbook = load_workbook(full_path)
        worksheet = workbook[report]
        worksheet = self.change_excel_colors(worksheet)
        worksheet = self.change_excel_column_sizes(worksheet)
        workbook.save(full_path)

        #if not os.path.exists(folder + filename):
        #df.to_excel(folder + filename, sheet_name = report, index = False)
    
    def change_excel_colors(self, worksheet):
        for row in worksheet.iter_rows():
            cell_color = self.colors['grey']
            # the order of this list determines which colors
            # override others
            for color in [self.time_based_color(row,worksheet),
            self.determine_resolved_color(row,worksheet,'Date Resolved','green'),
            self.determine_resolved_color(row,worksheet,'Manually Resolved','blue')]:
                if color != None:
                    cell_color = color
            for cell in row:
                cell.border = self.thin_border
                header_value = worksheet.cell(row=1,
                column=cell.column).value
                if header_value in ['Flag Count','Flags','Form']:
                    cell.fill = cell_color
                else:
                    cell.fill = self.colors['grey']
    
        return worksheet
    
    def time_based_color(self, excel_row, worksheet):
        for cell in excel_row:
            header_value = worksheet.cell(row=1, column=cell.column).value
            cell_val = str(cell.value)   
            if cell_val == 'None':
                cell_val = ''
            if header_value == 'Days Since Detected':
                if self.utils.can_be_float(cell_val):
                    days_since_detected = int(cell_val)
                    if days_since_detected < 7:
                        return self.colors['yellow']
                    elif 7 <= days_since_detected < 14:
                        return self.colors['orange']
                    else:
                        return self.colors['red']
        return None
    
    def determine_resolved_color(self, excel_row, worksheet, col_to_check, color_to_return):
        for cell in excel_row:
            cell.fill = self.colors['grey']
            header_value = worksheet.cell(row=1, column=cell.column).value
            cell_val = str(cell.value)   
            if cell_val == 'None':
                cell_val = ''
            if header_value == col_to_check:
                if cell_val !='' and cell_val != header_value:
                    color = self.colors[color_to_return]
                    return color

        return None

    def change_excel_column_sizes(self,worksheet):
        columns_sizes = {
            'Participant' : 10,
            'Timepoint' : 10,
            'Flag Count' : 10,
            'Form' : 35,
            'Flags' : 35,
            'Translations' : 35,
            'Days Since Detected' : 25,
            'Date Resolved': 20,
            'Manually Resolved' : 20,
            'Comments' : 15 
        }
        for header, length in columns_sizes.items():
            col_letter = self.find_col_letter(worksheet, header)
            if col_letter != None:  
                worksheet.column_dimensions[col_letter].width = length

        return worksheet

    def find_col_letter(self,worksheet, col_name):
        for cell in worksheet[1]:  
            if cell.value == col_name:
                column_letter = get_column_letter(cell.column)
                return column_letter
        return None
        
    def upload_files_to_dropbox(self):
        pass

    def merge_rows(self, values):
        unique_values  = list(dict.fromkeys(values))

        if len(unique_values) == 1:
            return values.iloc[0]
        else:
            return ' | '.join(unique_values)

    def convert_to_shared_format(self, raw_df, network):
        columns_names = self.formatted_column_names[network]
        columns_to_match = ['subject','displayed_timepoint','displayed_form']
        raw_df['date_resolved'] = ''
        raw_df.loc[raw_df['currently_resolved'] == True, 'date_resolved'] = (
            raw_df.loc[raw_df['currently_resolved'] == True, 'dates_resolved']
            .apply(lambda x: x.split(' | ')[-1])
        )

        merged_df = raw_df.groupby(columns_to_match).agg(self.merge_rows).reset_index()
        merged_df['flag_count'] = merged_df['error_message'].str.count(r'\|') + 1
        #merged_df['displayed_form'] = merged_df['displayed_form'].str.title().str.replace('_',' ')
        merged_df.rename(columns=columns_names, inplace=True)
        merged_df = merged_df[list(columns_names.values())]
        #move manually resolved to the bottom
        merged_df = self.move_rows_to_bottom('Manually Resolved',None, merged_df)
        merged_df = self.move_rows_to_bottom('Date Resolved','Manually Resolved', merged_df)
        
        merged_df.to_csv(f'{self.formatted_outputs_path}AMPSCZ_Output.csv',index = False)

        return merged_df
    
    def move_rows_to_bottom(self, incl_col_name,excl_col_name, df):
        if excl_col_name != None:
            moving_df = df[(df[incl_col_name] != '') & (df[excl_col_name]=='')]
            df = df[(df[incl_col_name] == '') | (df[excl_col_name]!='')]
        else:
            moving_df = df[df[incl_col_name] != '']
            df = df[df[incl_col_name] == '']

        result = pd.concat([df, moving_df], ignore_index=True)
        return result

    def save_to_dropbox(self, fullpath, local_path):
        dbx = self.utils.collect_dropbox_credentials()
        dropbox_path = f'/Apps/Automated QC Trackers/refactoring_tests/'
        with open(fullpath, 'rb') as f:
            dbx.files_upload(f.read(), dropbox_path + local_path,\
            mode=dropbox.files.WriteMode.overwrite)


