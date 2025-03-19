import os,sys
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
import pandas as pd,numbers,math,sys,openpyxl,datetime,random,matplotlib.pyplot as plt,collections,os,dropbox,re
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, colors, Protection,Color
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import range_boundaries,get_column_letter
from openpyxl import load_workbook,Workbook
import numpy as np
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import openpyxl
import dropbox
import webbrowser
import base64
import requests
import json
from io import BytesIO
import ast
from openpyxl.formatting.formatting import ConditionalFormattingList

from copy import copy

class GenerateTrackers():
    """Class to generate combined trackers
    for each network, then split up the trackers
    for specific sites"""

    def __init__(self,dataframe, site_name, sheet_title):
        self.location = 'pnl_server'
        self.test_prefix = 'Tests/' #Tests/
        if self.location =='pnl_server':
            self.combined_df_folder = '/data/predict1/data_from_nda/formqc/'
            self.combined_cognition_folder = ''
            self.penn_path = '/data/predict1/data_from_nda/'
            self.absolute_path = '/PHShome/ob001/anaconda3/new_forms_qc/QC/'
        else:
            self.combined_df_folder = ''
            self.penn_path = ''
            self.combined_cognition_folder = 'cognition/'
            self.absolute_path = 'C:/formqc/AMPSCZ_QC_and_Visualization/QC/'

        self.dataframe = dataframe
        self.sheet_title = sheet_title 
        self.site_name = site_name

        self.green_fill = PatternFill(start_color='b9d9b4', end_color='b9d9b4', fill_type='gray125')
        self.grey_fill = PatternFill(start_color='ededed', end_color='ededed', fill_type='gray125')
        self.white_fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='gray125')
        self.yellow_fill = PatternFill(start_color='bab466', end_color='bab466', fill_type='gray125')
        self.blue_fill = PatternFill(start_color='d8cffc', end_color='d8cffc', fill_type='gray125') 
        self.thick_purple_border = Border(left=Side(style='thick', color='0e176b'),right=Side(style='thick', color='0e176b'),
        top=Side(style='thick', color='0e176b'),bottom=Side(style='thick', color='0e176b'))
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),bottom=Side(style='thin'))
        self.all_pronet_sites = ["KC", "BI", "SD", "NL", "OR", "CA", "IR", "MU","YA", "HA",\
        "MA", "PI", "PV", "MT", "SF", "NC",'NN','PA','WU',"LA",'GA','TE','CM','SL','SI','SH','UR','OH']
        self.all_prescient_sites = ['BM', 'CG', 'CP', 'GW', 'HK', 'JE', 'LS', 'ME', 'SG', 'ST']
        self.site_full_name_translations = {'BI': 'Beth Israel (Harvard) (BI)',\
                'CA': 'Calgary, CA (CA)', 'CM': 'Cambridge (CM)', 'GA': 'Georgia (GA)',\
                'HA': 'Hartford (Institute of Living) (HA)', 'IR': 'UC Irvine (IR)',\
                'KC': "King's College, UK (KC)", 'LA': 'UCLA (LA)', 'MA': 'Madrid, Spain (MA)',\
                'MT': 'Montreal, CA (MT)', 'MU': 'Munich, Germany (MU)', 'NC': 'UNC (North Carolina) (NC)',\
                'NL': 'Northwell (NL)', 'NN': 'Northwestern (NN)', 'OR': 'Oregon (OR)',
                'PA': 'University of Pennsylvania (PA)', 'PI': 'Pittsburgh (UPMC) (PI)',\
                'PV': 'Pavia, Italy (PV)', 'SD': 'UCSD (SD)', 'SF': 'UCSF (Mission Bay) (SF)',\
                'SH': 'Shanghai, China (SH)', 'SI': 'Mt. Sinai (SI)', 'SL': 'Seoul, South Korea (SL)',\
                'TE': 'Temple (TE)', 'WU': 'Washington University (WU)', 'YA': 'Yale (YA)','UR':'University of Rochester (UR)',\
                'OH':'Ohio (OH)', 'BM': 'Birmingham, UK (BM)', 'CG': 'Cologne, DE (CG)', \
                'CP': 'Copenhagen, DK (CP)', 'GW': 'Gwangju, KR (GW)', 'HK': 'Hong Kong (HK)',\
                'JE': 'Jena, DE (JE)', 'LS': 'Lausanne, CH (LS)', 'ME': 'Melbourne (ME)',\
                'SG': 'Singapore (SG)', 'ST': 'Santiago (ST)',
                'PRONET':'PRONET','PRESCIENT':'PRESCIENT','AMPSCZ':'AMPSCZ'}

        self.prescient_site_col_translations = {'Subject':'ID',\
        'Timepoint': 'Timepoint of Error','Additional Comments':'Site Comments',\
        "Subject's Current Timepoint": "Participant's Current Timepoint"}

        self.new_prescient_order = ["Subject","Timepoint",\
            "Subject's Current Timepoint","General Flag",\
            "Specific Flags","Variable Translations","Sent to Site","Additional Comments",\
            "Manually Marked as Resolved",\
            "Date Resolved","Date Flag Detected","Time since Flag Detected"]

    def run_script(self):
        self.create_site_folders(True)
        self.create_ra_folders()
        self.save_to_excel(self.dataframe,self.sheet_title)
        self.create_site_folders()

    def reverse_dictionary(self,orig_dictionary):
        new_dictionary = {}
        for key,value in orig_dictionary.items():
            new_dictionary[value] = key
        return new_dictionary            

    def synchronize_resolved_rows(self,sheet):
        team_forms = {'Blood Report':['blood_sample_preanalytic_quality_assurance',\
        'cbc_with_differential'],\
        'Cognition Report':['penncnb','premorbid_iq_reading_accuracy',\
        'iq_assessment_wasiii_wiscv_waisiv'],
        'Scid Report':['scid5_psychosis_mood_substance_abuse']}
        for combined_sheet in ['Main Report','Secondary Report']:
            if sheet in ['Cognition Report','Blood Report']:
                rows_to_move = []
                combined_path = (f'{self.absolute_path}'
                f'site_outputs/{self.test_prefix}{self.site_name}/combined/{self.site_name}_Output.xlsx')
                if os.path.exists(combined_path):
                    combined_df = pd.read_excel(combined_path,\
                    keep_default_na=False,sheet_name=combined_sheet)
                    merged_df = pd.merge(self.dataframe, combined_df, 
                            on=['General Flag', 'Subject',\
                           'Timepoint','Specific Flags'], 
                            how='outer', 
                            suffixes=('', '_new'))
                    for index,row in merged_df.iterrows():
                        for form in team_forms[sheet]:
                            if form in row['General Flag'].replace(' ','_').lower()\
                            and str(row['Date Resolved']) == 'nan' and str(row['Date Resolved_new']) not in ['','nan']:
                                row['Date Resolved'] = row['Date Resolved_new']
                                row["Subject's Current Timepoint"] = row["Subject's Current Timepoint_new"] 
                                processed_row = {key: value for key, value in row.items() if not key.endswith('_new')}
                                rows_to_move.append(processed_row)

                    self.dataframe = self.dataframe.append(rows_to_move, ignore_index=True)


    def synchronize_dates(self,sheet):
        for combined_sheet in ['Main Report','Secondary Report']:
            if sheet in ['Cognition Report','Blood Report',\
            'Minor Data Missing','Substantial Data Missing','NDA Errors']:
                combined_path = (f'{self.absolute_path}'
                f'site_outputs/{self.test_prefix}{self.site_name}/combined/{self.site_name}_Output.xlsx')
                if os.path.exists(combined_path):
                    combined_df = pd.read_excel(combined_path,\
                    keep_default_na=False,sheet_name=combined_sheet)
                    merged_df = pd.merge(self.dataframe, combined_df, 
                            on=['General Flag', 'Subject', 'Timepoint','Specific Flags'], 
                            how='left', 
                            suffixes=('', '_new'))
                    for sync_col in ['Date Flag Detected',\
                    'Time since Flag Detected','Manually Marked as Resolved']:
                        self.dataframe[f'{sync_col}'] = merged_df[f'{sync_col}_new'].where(\
                        pd.notnull(merged_df[f'{sync_col}_new']), self.dataframe[f'{sync_col}'])

                    non_blank_date_resolved = (pd.notnull(\
                    self.dataframe['Date Resolved'])) & (self.dataframe['Date Resolved'] != '')

                    valid_updates_mask = (pd.notnull(\
                    merged_df['Date Resolved_new'])) & (merged_df['Date Resolved_new'] != '')
                    update_mask = non_blank_date_resolved & valid_updates_mask

                    self.dataframe.loc[update_mask, 'Date Resolved'] = merged_df.loc[update_mask, 'Date Resolved_new']
                    
                    """self.dataframe.loc[non_blank_date_resolved, 'Date Resolved'] = \
                    merged_df.loc[non_blank_date_resolved, f'Date Resolved'].where(
                        pd.notnull(merged_df.loc[non_blank_date_resolved, f'Date Resolved_new']),
                        merged_df.loc[non_blank_date_resolved, f'Date Resolved_new']
                    )"""
                    
    def compare_dataframes(self,sheet,filename):
        """Compares the new errors to the errors from the
        previous output in order to determine what has been
        resolved.

        Parameters
        --------------
        sheet: current sheet being checked
        filename: name of file being checked
        """

        old_dataframe =  pd.read_excel(filename, sheet_name=sheet)
        if self.site_name == 'PRESCIENT':
            if sheet == 'Main Report':
                old_df_filepath  = (f'{self.absolute_path}site_outputs/{self.test_prefix}'
                    f'{self.site_name}/combined/{self.site_name}_Output.xlsx')
                dropbox_filepath = (f'/Apps/Automated QC Trackers/'
                f'{self.test_prefix}PRESCIENT/')
                prescient_full_names = []
                for site in self.all_prescient_sites:
                    prescient_full_names.append(self.site_full_name_translations[site])
                old_dataframe = self.read_prescient_site_columns(old_dataframe,prescient_full_names,dropbox_filepath)
                dropbox_filepath = (f'/Apps/Automated QC Trackers/'
                f'{self.test_prefix}PRESCIENT/Melbourne (ME)/')
                ra_list = self.create_ra_list()
                old_dataframe = self.read_prescient_site_columns(old_dataframe,ra_list,dropbox_filepath)
            new_order = ["Subject","Timepoint",\
            "Subject's Current Timepoint","General Flag",\
            "Specific Flags","Variable Translations","Sent to Site","Additional Comments",\
            "Manually Marked as Resolved",\
            "Date Resolved","Date Flag Detected","Time since Flag Detected"]

            old_dataframe = old_dataframe[self.new_prescient_order]

            self.dataframe = self.dataframe[self.new_prescient_order]

        columns_to_keep = ['Manually Marked as Resolved',\
        'Additional Comments','Date Flag Detected']
            
        for i, old_df_row in old_dataframe.iterrows():
            match = False
            for j, new_df_row in self.dataframe.iterrows():
                if (old_df_row['Subject'] == new_df_row['Subject']) and\
                (old_df_row['Timepoint'] == new_df_row['Timepoint']) and\
                (old_df_row['General Flag'].split(':')[0].replace(' ','')\
                == new_df_row['General Flag'].split(':')[0].replace(' ','')):
                    #print(f"Match found in row {i} of df1 and row {j} of df2")
                    new_df_row['Date Resolved'] = ''
                    if old_df_row['Date Resolved'] != '':
                        specfic_flag_match = False
                        old_flags_list = old_df_row[\
                        'Specific Flags'].replace(' ','').replace('Eror','Error').split('|')
                        new_flags_list = new_df_row[\
                        'Specific Flags'].replace(' ','').replace('Eror','Error').split('|')
                        for flag in new_flags_list:
                            if flag in old_flags_list:
                                specfic_flag_match = True
                        if not specfic_flag_match:
                            old_df_row['Sent to Site'] = ''
                            old_df_row['Date Flag Detected'] = ''
                        old_df_row['Date Resolved'] = ''
                    for x in columns_to_keep:
                        new_df_row[x] = old_df_row[x]
                    match = True
            if match == False:
                if str(old_df_row['Date Resolved']) in ['nan','']:
                    old_df_row['Date Resolved'] = datetime.date.today()
                self.dataframe = self.dataframe.append(old_df_row, ignore_index=True)

    def move_datarame_rows(self):
        """Reorganizes rows to make it easier to
        separate resolved from unresolved errors"""

        rows_to_move = []
        for index, row in self.dataframe.iterrows():
            if str(row['Sent to Site']) not in ['','nan']:
                rows_to_move.append(row)
                self.dataframe = self.dataframe.drop(index)
        self.dataframe = self.dataframe.append(rows_to_move, ignore_index=True)
        rows_to_move = []
        for index, row in self.dataframe.iterrows():
            if str(row['Manually Marked as Resolved']) not in ['','nan']:
                rows_to_move.append(row)
                self.dataframe = self.dataframe.drop(index)
        for index, row in self.dataframe.iterrows():
            if str(row['Date Resolved']) not in ['','nan']:
                rows_to_move.append(row)
                self.dataframe = self.dataframe.drop(index)

        self.dataframe =\
        self.dataframe.append(rows_to_move, ignore_index=True)

    def save_to_excel(self,df,sheet,filename = '',
    backup_path = '', save_backup =True,compare = True):
        """Function to update excel sheet and account for 
        which files/sheets already exist

        Parameters
        ------------------
        sheet: sheet being saved
        filename: name of file being saved
        backup_path: path of backup file
        save_backup: whether or not backup will be saved
        """

        self.dataframe = df
        if sheet == 'Scid Report':
            report_str = ''
        else:
            report_str = ''
        if filename == '':
            filename =\
            (f'{self.absolute_path}site_outputs/{self.test_prefix}'
            f'{self.site_name}/combined/{self.site_name}_{report_str}Output.xlsx')
            backup_path =\
            (f'{self.absolute_path}Backups/{self.site_name}'
            f'_{report_str}Output_BACKUP_{datetime.date.today()}.xlsx')
        self.dataframe = self.dataframe.sort_values(by='Subject',\
        key=lambda x: x.str[:2]).sort_values(by='Subject')
        self.backup_path = backup_path
        self.save_backup = save_backup
        if os.path.exists(filename) and\
        self.sheet_title in pd.ExcelFile(filename).sheet_names:
            print(pd.ExcelFile(filename).sheet_names)
            self.compare_dataframes(sheet,filename)
            self.move_datarame_rows()
            self.synchronize_resolved_rows(sheet)
            self.synchronize_dates(sheet)

            if self.site_name == 'PRESCIENT':
                self.dataframe = self.dataframe[self.new_prescient_order]

            with pd.ExcelWriter(filename, mode='a',\
            engine='openpyxl',if_sheet_exists = 'replace') as writer:
                #writer.book = load_workbook(filename)
                self.dataframe.to_excel(writer, sheet_name=sheet, index=False)
            self.workbook = load_workbook(filename)
            worksheet = self.workbook[sheet]
            
        elif os.path.exists(filename):
            with pd.ExcelWriter(filename, mode='a',\
            engine='openpyxl',if_sheet_exists = 'replace') as writer:
                self.dataframe.to_excel(writer, sheet_name=sheet, index=False)
            self.workbook = load_workbook(filename)
            worksheet = self.workbook[sheet]
        else:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                self.dataframe.to_excel(writer, sheet_name=sheet, index=False)
            self.workbook = load_workbook(filename)
            worksheet = self.workbook[sheet]
    
        self.format_tracker(filename,sheet)

    def color_error_rows(self,cell,worksheet):
        """Determined color of row depending on 
        how long it has been since the error was detected

        Parameters
        ---------------------
        cell: current cell being modified
        worksheet: worksheet being modified
        """

        header_value = worksheet.cell(row=1, column=cell.column).value
        if header_value in self.cells_to_unlock:
            cell.protection = Protection(locked=False)
        if header_value == 'Manually Marked as Resolved':
            if cell.value.replace(' ','') != '':
                self.error_color = self.blue_fill
        elif header_value == 'Sent to Site':
            if cell.value != '':
                cell_border = self.thin_border
            else:
                cell_border = self.thin_border
        elif header_value == 'Date Flag Detected':
            self.time_since_flag = 0  
            if cell.value == '':
                cell.value = datetime.date.today()
            try:
                self.time_since_flag = datetime.date.today() -\
                datetime.datetime.strptime(str(cell.value).split(' ')[0],\
                "%Y-%m-%d").date()
            except Exception as e:
                print(e)
            if self.time_since_flag != 0:
                self.time_since_flag = self.time_since_flag.days

        elif header_value == 'Time since Flag Detected':
            if cell.value != header_value:
                if self.time_since_flag != 1:
                    cell.value = str(self.time_since_flag) + ' Days'
                else:
                    cell.value = str(self.time_since_flag) + ' Day'
            if self.time_since_flag == 0 or (\
            self.time_since_flag!= 0 and self.time_since_flag < 7):
                self.error_color = PatternFill(\
                start_color='e8e7b0', end_color='e8e7b0', fill_type='gray125')
            elif self.time_since_flag != 0 and 7 <= self.time_since_flag < 14:
                self.error_color = PatternFill(\
                start_color='ebba83', end_color='ebba83', fill_type='gray125')
            else:
                self.error_color = PatternFill(\
                start_color='de9590', end_color='de9590', fill_type='gray125')
        return cell

    def format_tracker(self,filename,sheet):
        """Adds additional formatting to output

        Parameters
        --------------
        filename: name of current file
        sheet: current sheet
        """

        self.workbook = load_workbook(filename)
        worksheet = self.workbook[sheet]
        worksheet.protection.sheet = True
        worksheet.protection.password = 'password_test'
        number_of_index_columns = 5
        columns_to_format = ['D', 'E']
        if self.site_name == 'PRESCIENT':
            column_to_check = 'I'
        else:
            column_to_check = 'J'
        rule = Rule(type="expression", dxf=DifferentialStyle(fill=self.blue_fill))
        formula =\
        f'ISBLANK(${get_column_letter(worksheet[column_to_check + "2"].column)}2)=FALSE'
        rule.formula = [formula]
        for column in columns_to_format:
            column_range = column + "2:" + column + str(worksheet.max_row)
            worksheet.conditional_formatting.add(column_range, rule)

        self.cells_to_unlock = ['Manually Marked as Resolved',\
        'Sent to Site','Additional Comments']

        self.color_resolved_rows(worksheet,filename)

    def color_resolved_rows(self,worksheet,filename):
        """Colors resolved rows green and 
        adds borders to rows that were 
        sent to site

        Parameters
        --------------
        worksheet: current sheet
        filename: current file 
        """

        for row in worksheet.iter_rows():
            self.time_since_flag = 0  
            for cell in row:
                if not isinstance(cell.value, str):
                    cell.value = str(cell.value)   
                if cell.value == 'None':
                    cell.value = ''
            compare_errors_dict = {'subject':'','Timepoint':'',\
            'form':'','specific_errors':''}
            for cell in row:  
                cell = self.color_error_rows(cell,worksheet)
            for cell in row:
                header_value = worksheet.cell(row=1, column=cell.column).value
                if header_value == 'Subject':
                    excel_subject = cell.value
                elif header_value == 'Timepoint':
                    excel_timepoint = cell.value
                cell.border = self.thin_border
                header_value = worksheet.cell(row=1, column=cell.column).value
                if header_value == 'Date Resolved' and cell.value != '':
                    self.error_color = self.green_fill 
            for cell in row:
                header_value = worksheet.cell(row=1, column=cell.column).value
                try:
                    if header_value == "General Flag" or\
                    header_value == "Specific Flags":
                        cell.fill = self.error_color
                        cell.border = cell_border
                    elif header_value in ["Additional Comments",\
                    "Manually Marked as Resolved"]:
                        cell.fill = self.white_fill    
                    else:       
                        cell.fill = self.grey_fill   
                except Exception as e:
                    continue     
        for col in range(1, worksheet.max_column + 1):
            worksheet.cell(row=1, column=col).fill = self.grey_fill
        self.adjust_column_length(worksheet,filename)


    def adjust_column_length(self,worksheet,filename):
        """Adjusts column length and freezes
        the first two columns

        Parameters
        --------------
        worksheet: current sheet
        filename: current file
        """
        

        for column in worksheet.columns:
            max_length = 0
            length_limit = 300
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length and\
                    len(str(cell.value)) < length_limit:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            if self.site_name == 'PRESCIENT':
                adjusted_width = (max_length + 60) * .2
            else:
                adjusted_width = (max_length + 110) * .2
            worksheet.column_dimensions[column_letter].width = adjusted_width
            
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=False,\
                shrink_to_fit=False)
        font = Font(size=16, bold=True)
        for cell in worksheet[1]:
            cell.font = font
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        worksheet.row_dimensions[1].height = 80
        color = worksheet['A1'].fill.start_color.index
        worksheet.freeze_panes = 'C2'
        columns_to_lock = []
        worksheet.print_area  = 'A1:N' + str(len(self.dataframe) + 50)
   
        self.workbook.save(filename)
        wb = load_workbook(filename=filename)
        wb_copy = load_workbook(filename=filename)
        if self.save_backup:
            wb_copy.save(self.backup_path)
        wb.close()

    def collect_dropbox_credentials(self):
        """reads dropbox credentials from
        JSON file"""

        with open(f'{self.absolute_path}dropbox_credentials.json', 'r') as file:
            json_data = json.load(file)
        APP_KEY = json_data['app_key']
        APP_SECRET = json_data['app_secret']
        REFRESH_TOKEN = json_data['refresh_token']

        dbx = dropbox.Dropbox(
            app_key = APP_KEY,
            app_secret = APP_SECRET,
            oauth2_refresh_token = REFRESH_TOKEN
        )

        return dbx

    def create_site_folders(self,download_combined = False):
        """Function to save files onto dropbox and create 
        individual site folders

        Parameters
        --------------------
        download_combined: determines if combined 
        copy will be first downloaded from dropbox
        """

        dbx = self.collect_dropbox_credentials()

        if download_combined:
            for network in ['PRONET','PRESCIENT']:
                dropbox_path = f'/Apps/Automated QC Trackers/{self.test_prefix}{network}/combined'
                for entry in dbx.files_list_folder(dropbox_path).entries:
                    print(entry.name)
                    if entry.name != f'{network}_Output.xlsx':
                        continue
                    _, res = dbx.files_download(dropbox_path + f'/{entry.name}')
                    data = res.content
                    with open(f'{self.absolute_path}site_outputs/{self.test_prefix}{network}/combined/{entry.name}', "wb") as f:
                        f.write(data)
                        print(f'{entry} data written')
        else:
            network = self.site_name
            dropbox_path = f'/Apps/Automated QC Trackers/{self.test_prefix}{network}/combined'
            for entry in dbx.files_list_folder(dropbox_path).entries:
                print(entry.name)
                if entry.name != f'{network}_Output.xlsx':
                    continue
                combined_path = f'{self.absolute_path}site_outputs/{self.test_prefix}{network}/combined/{entry.name}'
                with open(combined_path, 'rb') as f:
                    dbx.files_upload(f.read(),dropbox_path + f'/{entry.name}',\
                    mode=dropbox.files.WriteMode.overwrite)
                _, res = dbx.files_download(dropbox_path + f'/{entry.name}')
                data = res.content
                self.save_team_folders(combined_path,dbx,network)
                for sheet in ['Main Report']:                
                    if ('SCID' in entry.name and sheet != 'Scid Report')\
                    or ('SCID' not in entry.name and sheet == 'Scid Report'):
                        continue
                    combined_df = pd.read_excel(BytesIO(data),\
                    sheet_name=sheet, keep_default_na = False)
                    if network == 'PRONET':
                        site_list = self.all_pronet_sites
                    else:
                        site_list = self.all_prescient_sites
                    for site in site_list:
                        print(site)
                        site_full_name = self.site_full_name_translations[site]
                        site_folder = f'{self.absolute_path}site_outputs/{self.test_prefix}{network}/{site_full_name}'
                        site_path = site_folder + f'/{entry.name}'
                        if self.sheet_title == 'Main Report':
                            self.save_site_output(combined_path,site,site_path,site_folder)
                            dropbox_path =  (f'/Apps/Automated QC Trackers/'
                                             f'{self.test_prefix}{network}/{site_full_name}/{entry.name}')
                            print(dropbox_path)
                            print(site_path)
                            print(dbx)
                            self.upload_subfolder(site_path, dropbox_path, dbx)
                    if self.site_name == 'PRESCIENT':
                        ra_list = self.create_ra_list()
                        with open(f'{self.absolute_path}ra_assignments.json', 'r') as file:
                            ra_assignments = json.load(file)
                        ra_assignments = {key.replace(" ", "_"): value for key, value in ra_assignments.items()}
                        print(ra_assignments)
                        for ra in ra_list:
                            print(ra)
                            ra_folder = (f'{self.absolute_path}site_outputs/'
                            f'{self.test_prefix}{network}/Melbourne (ME)/{ra}')
                            ra_path = ra_folder + f'/{entry.name}'
                            print(ra_path)
                            if self.sheet_title == 'Main Report':
                                self.save_site_output(combined_path,'ME', ra_path, ra_folder, ra_assignments[ra])
                                dropbox_path = (f'/Apps/Automated QC Trackers/'
                                f'{self.test_prefix}{network}/Melbourne (ME)/{ra}/{entry.name}')
                                self.upload_subfolder( ra_path, dropbox_path, dbx)
                                    
    def upload_subfolder(self, full_path,dropbox_path, dbx):
        self.remove_output_columns(full_path)
        if self.site_name =='PRESCIENT':
            self.rename_excel_columns(full_path, self.prescient_site_col_translations)
        with open(full_path, 'rb') as f: 
            dbx.files_upload(f.read(),\
            dropbox_path,\
            mode=dropbox.files.WriteMode.overwrite)

    def remove_output_columns(self,path):
        with open(path,'rb') as f:
            wb = load_workbook(path)
            ws = wb['Main Report']  
            column_to_remove = self.get_column_number(ws,'Sent to Site')
            ws.delete_cols(column_to_remove)
            wb.save(path)

    def get_column_number(self, sheet, column_name, header_row=1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=header_row, column=col).value == column_name:
                return col
        return None

    def rename_excel_columns(self,path, translations):
        with open(path,'rb') as f:
            wb = load_workbook(path)
            ws = wb['Main Report']  
        for old_name, new_name in translations.items():
            col_num = self.get_column_number(ws, old_name)
            if ws.cell(row=1, column=col_num).value == old_name:
                ws.cell(row=1, column=col_num).value = new_name
        wb.save(path)

    def create_ra_list(self):
        with open(f'{self.absolute_path}ra_assignments.json', 'r') as file:
            ra_assignments = json.load(file)
        print(ra_assignments)
        ra_list = list(ra_assignments.keys())   
        ra_list = [ra.replace(' ','_') for ra in ra_list]
        return ra_list

    def create_ra_folders(self):
        site_folder = (f'{self.absolute_path}site_outputs/'
        f'{self.test_prefix}PRESCIENT/{self.site_full_name_translations["ME"]}')
        ra_list = self.create_ra_list()
        
        for ra in ra_list:
            ra_path = site_folder + '/' + ra
            print(ra_path)
            if not os.path.exists(ra_path):
                os.makedirs(ra_path)
        print(ra_list)

    def read_prescient_site_columns(self, parent_df, child_df_list, dropbox_parent_path):
        dbx = self.collect_dropbox_credentials()
        def check_dropbox_path_exists(path):
            try:
                dbx.files_get_metadata(path)
                return True
            except dropbox.exceptions.ApiError as e:
                if isinstance(e.error, dropbox.files.GetMetadataError):
                    return False
                else:
                    raise e

        combined_df = parent_df
        for subcategory in child_df_list:
            print(subcategory)
            dropbox_path = (f'{dropbox_parent_path}{subcategory}')
            if not check_dropbox_path_exists(dropbox_path):
                continue
            for entry in dbx.files_list_folder(dropbox_path).entries:
                print(entry.name)
                if entry.name != f'PRESCIENT_Output.xlsx':
                    continue
                _, res = dbx.files_download(dropbox_path + f'/{entry.name}')
                data = res.content
                site_df = pd.read_excel(BytesIO(data),\
                sheet_name='Main Report')

                reverse_translations = self.reverse_dictionary(self.prescient_site_col_translations)
                site_df.rename(columns=reverse_translations, inplace=True)

                site_df.columns = site_df.columns.str.replace(' ','_')
                combined_df.columns = combined_df.columns.str.replace(' ','_')
                print(site_df['Manually_Marked_as_Resolved'])
            
                combined_df = pd.merge(combined_df, site_df, on=['Subject', 'Timepoint', 'General_Flag'], how='left', suffixes=('', '_site'))

                combined_df['Manually_Marked_as_Resolved'] = combined_df['Manually_Marked_as_Resolved'].fillna(combined_df['Manually_Marked_as_Resolved_site'])
                combined_df['Additional_Comments'] = combined_df['Additional_Comments'].fillna(combined_df['Additional_Comments_site'])

                combined_df['Manually_Marked_as_Resolved'] = combined_df['Manually_Marked_as_Resolved'].fillna(combined_df['Manually_Marked_as_Resolved'])
                combined_df['Additional_Comments'] = combined_df['Additional_Comments'].fillna(combined_df['Additional_Comments'])
                combined_df = combined_df.loc[:, ~combined_df.columns.str.endswith('_site')]
                print(combined_df['Manually_Marked_as_Resolved'])

                combined_df.columns = combined_df.columns.str.replace('_',' ')
                
                #combined_df.to_excel(filepath, index = False)

        return combined_df


    def save_site_output(self,combined_path,site_prefix,site_path,site_folder, ra_list = []):
        """Function to remove all subjects
        not in site and save it to the site folder

        Parameters
        --------------
        combined_path: path to combined tracker
        site_prefix: Site
        site_path: path to new site folder
        """

        if not os.path.exists(site_folder):
            os.makedirs(site_folder)

        wb = load_workbook(combined_path)
        ws = wb['Main Report']
        if self.site_name == 'PRESCIENT':
            ws.conditional_formatting = ConditionalFormattingList()
            column_to_check = 'H'
            rule = Rule(type="expression", dxf=DifferentialStyle(fill=self.blue_fill))
            formula =\
            f'ISBLANK(${get_column_letter(ws[column_to_check + "2"].column)}2)=FALSE'
            rule.formula = [formula]
            columns_to_format = ['D', 'E']
            for column in columns_to_format:
                column_range = column + "2:" + column + str(ws.max_row)
                ws.conditional_formatting.add(column_range, rule)

        def remove_rows(ws, ra_list):
            print(ra_list)
            check_column_letter = 'A'
            check_column_index\
            = openpyxl.utils.column_index_from_string(check_column_letter)
            rows_to_remove = []
            for row in ws.iter_rows(min_row=2):  
                cell_value = str(row[check_column_index - 1].value)
                if (not cell_value.startswith(site_prefix))\
                or (cell_value not in ra_list and ra_list != []):
                    rows_to_remove.append(row[0].row)
            count = 0
            start_index = None
            for current_index in reversed(rows_to_remove):
                if start_index is None:
                    start_index = current_index
                    count = 1
                elif current_index == start_index - 1:
                    start_index -= 1
                    count += 1
                else:
                    ws.delete_rows(start_index, count)
                    start_index = current_index
                    count = 1
            if start_index is not None:
                ws.delete_rows(start_index, count)
            return ws
        
        sheets_to_keep = ['Main Report', 'Manual Monitoring']
        for sheet in sheets_to_keep:
            ws = wb[sheet]
            ws = remove_rows(ws, ra_list)
        sheets_to_remove = [sheet.title for sheet in wb \
        if sheet.title not in sheets_to_keep]
        for sheet_name in sheets_to_remove:
            std = wb[sheet_name]
            wb.remove(std)

        wb.save(site_path)

    def save_team_folders(self,combined_path,dbx,network):
        wb = load_workbook(combined_path)
        report_folder_translations = {'Blood Report':'Blood',\
        'Cognition Report':'Cognition','Scid Report':'Scid'}
        for team_report, team_folder in report_folder_translations.items():
            if team_report not in pd.ExcelFile(combined_path).sheet_names:
                continue
            source_sheet = wb[team_report] 
            new_workbook = Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = source_sheet.title
            for row in source_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
            for col_letter, column_dimension in source_sheet.column_dimensions.items():
                    new_sheet.column_dimensions[col_letter].width = column_dimension.width

            new_workbook.save(f'{self.absolute_path}site_outputs/{self.test_prefix}{network}/Teams/{team_folder}/{network}_Output.xlsx')
            with open(f'{self.absolute_path}site_outputs/{self.test_prefix}{network}/Teams/{team_folder}/{network}_Output.xlsx', 'rb') as f:
                dbx.files_upload(f.read(),\
                f'/Apps/Automated QC Trackers/{self.test_prefix}{network}/Teams/{team_folder}/{network}_Output.xlsx',\
                mode=dropbox.files.WriteMode.overwrite)


#GenerateTrackers('', 'PRESCIENT', 'Main Report').create_site_folders()
